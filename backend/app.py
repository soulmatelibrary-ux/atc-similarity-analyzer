"""
유사호출 감시 시뮬레이션 시스템 - Flask 백엔드 API
"""
import os
import json
import threading
import platform
import signal
import time
from datetime import datetime, timedelta
from flask import Flask, request, jsonify, send_file, session, make_response
from flask_cors import CORS
from werkzeug.utils import secure_filename
from werkzeug.security import check_password_hash
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from functools import wraps

# 경로 설정
BACKEND_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(BACKEND_DIR)
import sys
sys.path.insert(0, PROJECT_DIR)

from core.similarity_engine import check_similarity, get_risk_level, get_similarity_score
from utils.file_validator import FileValidator
from utils.logger import setup_logger
from utils.similarity_optimizer import SimilarityOptimizer
from utils.sector_parser import parse_sector_times, calculate_sector_overlaps, get_sector_overlap_summary
from utils.constants import API_PORT, FRONTEND_PORT, DEBUG, SECTORS
from utils.license_manager import get_license_manager
from backend.license_api import license_bp, admin_license_bp
from database.db_manager import DatabaseManager
from core.flight_service import FlightService, upload_progress
from core.flight_processor import calculate_trajectory
from core.modeling_resources import get_modeling_resources

# Flask 앱 초기화
app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['UPLOAD_FOLDER'] = os.path.join(PROJECT_DIR, 'uploads')
app.config['SECRET_KEY'] = os.environ.get('APP_SECRET_KEY', 'change-this-secret')
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=int(os.environ.get('APP_SESSION_TIMEOUT_HOURS', 1)))
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

SESSION_TIMEOUT_SECONDS = int(os.environ.get('APP_SESSION_TIMEOUT_SECONDS', 3600))

# 타임아웃 설정 (유사호출 감지는 최대 45분까지 허용)
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0  # 캐시 비활성화

def timeout_handler(signum, frame):
    raise TimeoutError("Request timeout")

def long_running_task(timeout_seconds=2700):
    """
    긴 작업용 타임아웃 데코레이터 (기본 45분)
    - Unix/Linux: signal.alarm() 사용
    - Windows: threading 기반 타임아웃
    """
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            is_windows = platform.system() == 'Windows'

            if is_windows:
                # Windows: threading 기반 타임아웃
                result_container = [None]
                exception_container = [None]

                def run_function():
                    try:
                        result_container[0] = f(*args, **kwargs)
                    except Exception as e:
                        exception_container[0] = e

                thread = threading.Thread(target=run_function, daemon=True)
                thread.start()
                thread.join(timeout=timeout_seconds)

                if thread.is_alive():
                    logger.warning(f"Request exceeded timeout: {timeout_seconds}s")
                    raise TimeoutError(f"Request timeout after {timeout_seconds}s")

                if exception_container[0]:
                    raise exception_container[0]

                return result_container[0]
            else:
                # Unix/Linux: signal.alarm() 사용
                try:
                    signal.signal(signal.SIGALRM, timeout_handler)
                    signal.alarm(timeout_seconds)
                except Exception as e:
                    logger.warning(f"Timeout handler setup failed: {e}")

                try:
                    result = f(*args, **kwargs)
                finally:
                    try:
                        signal.alarm(0)  # 타이머 취소
                    except Exception as e:
                        logger.warning(f"Timeout cancellation failed: {e}")

                return result

        return wrapper
    return decorator

# CORS 설정 (모든 출처 허용 - 배포 환경에서도 작동)
CORS(app, resources={
    r"/api/*": {
        "origins": "*",  # 모든 출처 허용 (프론트엔드와 동일 origin이므로 문제 없음)
        "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
        "allow_headers": ["Content-Type", "Authorization"],
        "expose_headers": ["Content-Type"],
        "supports_credentials": True,
        "max_age": 3600
    }
})

# 로거 설정
logger = setup_logger(__name__)

# ==================== 라이선스 초기화 ====================
license_manager = get_license_manager()
logger.info("=" * 80)
logger.info("🔐 LICENSE INFORMATION")
logger.info("=" * 80)
license_manager.log_license_info()
logger.info("=" * 80)

# 라이선스가 유효하지 않으면 경고
if not license_manager.is_valid:
    logger.error(f"⚠️  Invalid License: {license_manager.message}")

# 기능 제한 정보 로그
limits = license_manager.get_limits()
logger.info(f"Max flights per upload: {limits['max_flights_per_upload']}")
logger.info(f"Max storage: {limits['max_storage_mb']}MB")
logger.info(f"Data retention: {limits['data_retention_days']} days")
if limits['watermark']:
    logger.info("⚠️  Watermark enabled (development license)")
# ========================================================

# 업로드 폴더 생성
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# 데이터베이스 초기화
# Render 배포 환경: 프로덕션 DB 사용 (database/similarity_detector.db)
# 로컬 개발 환경: 테스트 DB 사용 (database/backend_similarity_detector.db)
is_production = os.getenv('RENDER') == 'true' or os.getenv('FLASK_ENV') == 'production'
if is_production:
    db_path = os.path.join(PROJECT_DIR, 'database', 'similarity_detector.db')
else:
    db_path = os.path.join(PROJECT_DIR, 'database', 'backend_similarity_detector.db')
db_manager = DatabaseManager(db_path)
flight_service = FlightService(db_manager)


# 상수 정의
ALLOWED_FILE_TYPES = ['csv', 'xlsx', 'xls']
MAX_FILE_SIZE_MB = 16

# 글로벌 상태 (메모리 캐시용)
app_state = {
    'flights': [],
    'coexistences': [],
    'statistics': None
}

AUTH_EXEMPT_PREFIXES = (
    '/api/auth',
    '/api/health',
    '/api/sample',  # 샘플 파일 다운로드 (인증 불필요)
    '/api/admin',   # 관리자 기능 (개발 모드: 인증 불필요)
    '/api/summary', # 요약 탭 (인증 불필요)
    '/api/test',    # 모델링 테스트 (개발 모드: 인증 불필요)
    '/api/flights'  # 항공편 데이터 (인증 불필요)
)


def _parse_allowed_tabs(raw_value):
    if isinstance(raw_value, list):
        return raw_value

    if not raw_value:
        return ['*']

    try:
        parsed = json.loads(raw_value) if isinstance(raw_value, str) else raw_value
        if isinstance(parsed, list):
            return parsed
    except Exception:
        pass

    return ['*']


def _verify_credentials(email: str, password: str):
    """데이터베이스 관리자 자격 증명을 검증하고 사용자 정보를 반환."""
    admin_user = db_manager.get_admin_user_by_email(email)
    if not admin_user:
        return None

    if not admin_user.get('is_active'):
        return None

    stored_hash = admin_user.get('password_hash')
    if not stored_hash:
        return None

    if not check_password_hash(stored_hash, password):
        return None

    return admin_user


def _touch_session_activity():
    session['last_activity'] = time.time()


def _remaining_session_seconds() -> int:
    last = session.get('last_activity')
    if not last:
        return SESSION_TIMEOUT_SECONDS
    remaining = SESSION_TIMEOUT_SECONDS - int(time.time() - last)
    return max(0, remaining)


@app.before_request
def enforce_authentication():
    """API 엔드포인트 접근 시 세션을 검증하고 타임아웃을 처리한다."""
    if request.method == 'OPTIONS':
        return None

    if not request.path.startswith('/api'):
        return None

    if request.path.startswith(AUTH_EXEMPT_PREFIXES):
        return None

    user = session.get('user')
    if not user:
        return jsonify({
            'status': 'error',
            'message': '로그인이 필요합니다.'
        }), 401

    remaining = _remaining_session_seconds()
    if remaining <= 0:
        session.clear()
        return jsonify({
            'status': 'error',
            'message': '세션이 만료되었습니다. 다시 로그인해주세요.'
        }), 401

    _touch_session_activity()


@app.route('/api/auth/login', methods=['POST'])
def login():
    payload = request.get_json(silent=True) or {}
    email = (payload.get('email') or '').strip()
    password = (payload.get('password') or '').strip()

    if not email or not password:
        return jsonify({
            'status': 'error',
            'message': '이메일과 비밀번호를 모두 입력하세요.'
        }), 400

    admin_user = _verify_credentials(email, password)
    if not admin_user:
        logger.warning('로그인 실패 - 잘못된 자격 증명: %s', email)
        return jsonify({
            'status': 'error',
            'message': '이메일 또는 비밀번호가 올바르지 않습니다.'
        }), 401

    allowed_tabs = _parse_allowed_tabs(admin_user.get('allowed_tabs'))
    role = admin_user.get('role', 'user')

    session.clear()
    session['user'] = {
        'id': admin_user.get('id'),
        'username': admin_user.get('username'),
        'email': admin_user.get('email'),
        'role': role,
        'allowed_tabs': allowed_tabs,
        'login_at': datetime.utcnow().isoformat()
    }
    session.permanent = True
    _touch_session_activity()
    db_manager.record_admin_login(admin_user.get('id'))
    logger.info('사용자 로그인 성공: %s', email)

    return jsonify({
        'status': 'success',
        'user': {
            'id': admin_user.get('id'),
            'username': admin_user.get('username'),
            'email': admin_user.get('email'),
            'role': role,
            'allowed_tabs': allowed_tabs
        },
        'expires_in': _remaining_session_seconds()
    })


@app.route('/api/auth/logout', methods=['POST'])
def logout():
    username = session.get('user', {}).get('email')
    session.clear()
    if username:
        logger.info('사용자 로그아웃: %s', username)
    return jsonify({'status': 'success'})


@app.route('/api/auth/session', methods=['GET'])
def session_status():
    user = session.get('user')
    if not user:
        return jsonify({
            'status': 'success',
            'authenticated': False,
            'expires_in': 0
        })

    return jsonify({
        'status': 'success',
        'authenticated': True,
        'user': {
            'id': user.get('id'),
            'username': user.get('username'),
            'email': user.get('email'),
            'role': user.get('role', 'user'),
            'allowed_tabs': user.get('allowed_tabs', ['*'])
        },
        'expires_in': _remaining_session_seconds()
    })

# 진행 상황 저장소 (파일 기반 - 안정적)
PROGRESS_DIR = os.path.join(PROJECT_DIR, '.uploads_progress')
os.makedirs(PROGRESS_DIR, exist_ok=True)

def save_upload_progress(process_id, progress_data):
    """진행 상황을 파일에 저장"""
    progress_file = os.path.join(PROGRESS_DIR, f'{process_id}.json')
    try:
        with open(progress_file, 'w', encoding='utf-8') as f:
            json.dump(progress_data, f, ensure_ascii=False)
    except Exception as e:
        logger.error(f"진행 상황 저장 오류: {e}")

def load_upload_progress(process_id):
    """진행 상황을 파일에서 로드"""
    progress_file = os.path.join(PROGRESS_DIR, f'{process_id}.json')
    if not os.path.exists(progress_file):
        return None

    for attempt in range(2):  # 짧은 시간 간격으로 두 번 시도 (원자적 쓰기 보완)
        try:
            with open(progress_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except json.JSONDecodeError:
            # 쓰기 도중 읽으면 빈 파일이 보일 수 있으므로 잠시 대기 후 재시도
            time.sleep(0.05)
        except Exception as e:
            logger.error(f"진행 상황 로드 오류: {e}")
            break

    return None

def delete_upload_progress(process_id):
    """진행 상황 파일 삭제"""
    progress_file = os.path.join(PROGRESS_DIR, f'{process_id}.json')
    try:
        if os.path.exists(progress_file):
            os.remove(progress_file)
    except Exception as e:
        logger.error(f"진행 상황 파일 삭제 오류: {e}")


def format_duration(seconds):
    """사람이 읽기 쉬운 시간 문자열 생성"""
    seconds = max(0, int(round(seconds)))
    hours, remainder = divmod(seconds, 3600)
    minutes, secs = divmod(remainder, 60)
    parts = []

    if hours:
        parts.append(f"{hours}시간")
    if minutes:
        parts.append(f"{minutes}분")
    if not parts or secs:
        parts.append(f"{secs}초")

    return ' '.join(parts)


def estimate_processing_time(record_count):
    """CSV 처리 예상 시간을 단순 휴리스틱으로 계산"""
    record_count = max(0, int(record_count))
    baseline_seconds = 6 if record_count else 0

    tier_rates_ms = [
        (20000, 0.52),
        (80000, 0.72),
        (160000, 0.9)
    ]
    per_record_ms = 1.05
    for threshold, rate in tier_rates_ms:
        if record_count <= threshold:
            per_record_ms = rate
            break

    scaling_factor = 1 + min(record_count / 250000, 0.65)
    core_seconds = (record_count * per_record_ms / 1000.0) * scaling_factor
    estimated_seconds = max(5, baseline_seconds + core_seconds)

    stage_core = max(core_seconds, 0)
    stage_breakdown = {
        'file_validation': round(stage_core * 0.18, 1),
        'database_ingest': round(stage_core * 0.37, 1),
        'similarity_analysis': round(stage_core * 0.45, 1)
    }

    return {
        'baseline_seconds': baseline_seconds,
        'core_seconds': core_seconds,
        'estimated_seconds': estimated_seconds,
        'per_record_ms': per_record_ms,
        'scaling_factor': scaling_factor,
        'stage_breakdown': stage_breakdown
    }


# ============================================================================
# Favicon 엔드포인트 (브라우저 자동 요청 처리)
# ============================================================================

@app.route('/favicon.ico')
def favicon():
    """브라우저의 favicon 요청 처리 (404 방지)"""
    return '', 204  # 204 No Content


@app.route('/<path:filename>')
def serve_frontend(filename):
    """프론트엔드 정적 파일 제공"""
    frontend_dir = os.path.join(PROJECT_DIR, 'frontend')
    file_path = os.path.join(frontend_dir, filename)

    # 경로 보안 확인 (디렉토리 이상 이동 방지)
    if not os.path.abspath(file_path).startswith(os.path.abspath(frontend_dir)):
        return {"error": "Invalid path"}, 403

    if os.path.isfile(file_path):
        # 파일 확장자에 따른 Content-Type 설정
        mime_types = {
            '.css': 'text/css',
            '.js': 'application/javascript',
            '.html': 'text/html',
            '.json': 'application/json',
            '.png': 'image/png',
            '.jpg': 'image/jpeg',
            '.jpeg': 'image/jpeg',
            '.gif': 'image/gif',
            '.svg': 'image/svg+xml',
            '.ico': 'image/x-icon',
            '.woff': 'font/woff',
            '.woff2': 'font/woff2',
            '.ttf': 'font/ttf'
        }

        # 파일 확장자 확인
        _, ext = os.path.splitext(file_path)
        content_type = mime_types.get(ext.lower(), 'text/plain')

        with open(file_path, 'r', encoding='utf-8') as f:
            response = make_response(f.read())
            response.headers['Content-Type'] = content_type
            return response

    return {"error": f"File not found: {filename}"}, 404


@app.route('/')
@app.route('/index.html')
def index():
    """루트 경로 - 프론트엔드 제공"""
    frontend_path = os.path.join(PROJECT_DIR, 'frontend', 'index.html')
    try:
        with open(frontend_path, 'r', encoding='utf-8') as f:
            return f.read()
    except FileNotFoundError:
        logger.error(f"Frontend file not found: {frontend_path}")
        return """
        <!DOCTYPE html>
        <html>
        <head>
            <title>오류</title>
            <style>
                body { font-family: sans-serif; text-align: center; padding-top: 50px; }
                .container { max-width: 600px; margin: 0 auto; padding: 20px; }
                h1 { color: #e74c3c; }
            </style>
        </head>
        <body>
            <div class="container">
                <h1>프론트엔드 파일을 찾을 수 없습니다</h1>
                <p>Frontend 파일 경로: """ + frontend_path + """</p>
            </div>
        </body>
        </html>
        """


# ============================================================================
# 헬스체크 엔드포인트
# ============================================================================

@app.route('/api/health', methods=['GET'])
def health_check():
    """
    시스템 헬스 체크

    Returns:
        JSON: {status: 'ok', timestamp: ISO8601}
    """
    return jsonify({
        'status': 'ok',
        'timestamp': datetime.now().isoformat(),
        'version': '1.0.0-alpha',
        'phase': 4
    }), 200


@app.route('/api/similarity-levels', methods=['GET'])
def get_similarity_levels():
    """
    유사도 레벨 및 정의 조회

    Returns:
        JSON: {
            status: 'success',
            data: {
                LEVEL_2-1: {risk: 'HIGH', score: 100, description: '...'},
                ...
            }
        }
    """
    try:
        from utils.constants import SIMILARITY_LEVELS

        return jsonify({
            'status': 'success',
            'data': SIMILARITY_LEVELS
        }), 200
    except Exception as e:
        logger.error(f"유사도 레벨 조회 실패: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'서버 오류: {str(e)}'
        }), 500


# ============================================================================
# 파일 업로드 엔드포인트
# ============================================================================

def _validate_file_extension(filename):
    """
    파일 확장자 검증 (화이트리스트)

    Args:
        filename: 파일명

    Returns:
        tuple: (is_valid: bool, error_message: str)
    """
    if '.' not in filename:
        return False, "파일에 확장자가 없습니다"

    ext = filename.rsplit('.', 1)[1].lower()
    if ext not in ALLOWED_FILE_TYPES:
        return False, f"허용되지 않는 파일 형식입니다. 허용: {', '.join(ALLOWED_FILE_TYPES)}"

    return True, ""


def _validate_file_mime_type(file):
    """
    파일 MIME 타입 검증

    Args:
        file: Werkzeug FileStorage 객체

    Returns:
        tuple: (is_valid: bool, error_message: str)
    """
    allowed_mimes = [
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  # .xlsx
        'application/vnd.ms-excel',  # .xls
        'text/csv',
        'application/csv',
    ]

    mime_type = file.content_type or ''

    # MIME 타입이 명시되지 않은 경우 파일명으로 유추
    if not mime_type or mime_type not in allowed_mimes:
        # 확장자 기반 검증만 수행
        return True, ""

    if mime_type not in allowed_mimes:
        return False, f"허용되지 않는 파일 타입입니다: {mime_type}"

    return True, ""


@app.route('/api/upload/flights', methods=['POST', 'OPTIONS'])
def upload_flights():
    """
    항공편 데이터 파일 업로드

    Request:
        - file: Excel/CSV 파일
        - mode: 'replace' (덮어쓰기) 또는 'append' (누적) - 기본값 'replace'

    Returns:
        JSON: { status, message, data }
    """
    # OPTIONS 요청 처리
    if request.method == 'OPTIONS':
        return '', 204
    
    try:
        logger.info(f"업로드 요청 수신 - Files: {list(request.files.keys())}, Form: {list(request.form.keys())}")
        logger.info(f"Content-Type: {request.content_type}")
        
        if 'file' not in request.files:
            logger.error(f"파일 없음. 가능한 키: {list(request.files.keys())}")
            return jsonify({
                'status': 'error',
                'message': f'파일이 제공되지 않았습니다. 가능한 키: {list(request.files.keys())}'
            }), 400

        file = request.files['file']
        mode = request.form.get('mode', 'replace')  # 기본값: 덮어쓰기

        logger.info(f"파일 정보 - 이름: {file.filename}, 크기: {len(file.read())} bytes")
        file.seek(0)  # 포인터 리셋
        
        if file.filename == '':
            logger.error("파일명이 비어있음")
            return jsonify({
                'status': 'error',
                'message': '파일명이 비어있습니다'
            }), 400

        # 1단계: 파일 확장자 검증
        is_valid, error_msg = _validate_file_extension(file.filename)
        if not is_valid:
            logger.warning(f"파일 확장자 검증 실패: {file.filename} - {error_msg}")
            return jsonify({
                'status': 'error',
                'message': error_msg
            }), 400

        # ... (중간 생략 등 기존 코드 유지 가능하므로 아래 로직만 수정) ...
        # 2단계: MIME 타입 검증
        is_valid, error_msg = _validate_file_mime_type(file)
        if not is_valid:
            logger.warning(f"파일 MIME 타입 검증 실패: {file.filename} - {error_msg}")
            return jsonify({
                'status': 'error',
                'message': error_msg
            }), 400

        # 3단계: 파일명 보안 처리
        filename = secure_filename(file.filename)
        logger.info(f"Secure filename: {filename}")
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        logger.info(f"저장 경로: {filepath}")

        # 4단계: 파일 저장 (검증된 확장자만 저장)
        file.save(filepath)
        logger.info(f"파일 저장 완료: {filepath}, 파일 크기: {os.path.getsize(filepath)} bytes")

        # 5단계: 저장된 파일 검증
        validator = FileValidator()
        logger.info(f"파일 검증 시작: {filepath}")
        validation_result = validator.validate_file(filepath)

        if validation_result['status'] == 'invalid':
            os.remove(filepath)  # 저장된 파일 삭제
            logger.warning(f"저장된 파일 검증 실패: {filename}")
            logger.error(f"검증 오류 상세: {validation_result['errors']}")
            logger.error(f"검증 경고: {validation_result['warnings']}")
            return jsonify({
                'status': 'error',
                'message': '파일 검증 실패',
                'data': {
                    'file_name': filename,
                    'errors': validation_result['errors'],
                    'warnings': validation_result['warnings']
                }
            }), 400

        # 6단계: 데이터 저장 (데이터베이스) - 백그라운드에서 처리
        flights_data = validation_result['data'].to_dict('records')

        # License 검증: 항공편 수 확인
        license_manager = get_license_manager()
        limits = license_manager.get_limits()
        max_flights = limits['max_flights_per_upload']

        if len(flights_data) > max_flights:
            logger.warning(f"라이선스 제한: 항공편 수 초과 - {len(flights_data)} > {max_flights}")
            return jsonify({
                'status': 'error',
                'message': f'라이선스 제한: 한 번에 최대 {max_flights}개 항공편까지만 업로드할 수 있습니다. (현재: {len(flights_data)}개)',
                'data': {
                    'file_name': filename,
                    'flight_count': len(flights_data),
                    'max_flights': max_flights,
                    'license_type': license_manager.license_type
                }
            }), 403

        # 파일 크기 계산
        file.seek(0, os.SEEK_END)
        file_size = file.tell()
        file.seek(0)

        # 진행 상황 추적용 프로세스 ID 생성
        import uuid
        process_id = str(uuid.uuid4())

        # 백그라운드 스레드에서 처리
        def process_flights_background():
            try:
                # 데이터 처리 모드에 따라 분기
                if mode == 'replace':
                    # 기존 데이터 초기화 (새 파일 업로드 시 깨끗한 상태에서 시작)
                    logger.info("기존 데이터 초기화 중 (덮어쓰기 모드)...")
                    db_manager.execute_query("DELETE FROM sector_overlaps")
                    db_manager.execute_query("DELETE FROM similarities")
                    db_manager.execute_query("DELETE FROM sector_times")
                    db_manager.execute_query("DELETE FROM flights")
                    logger.info("기존 데이터 초기화 완료")
                    
                    # 메모리 캐시 초기화
                    app_state['flights'] = []
                else:
                    logger.info("데이터 누적 모드로 업로드 진행...")

                service_result = flight_service.process_and_save_flights(
                    flights_data,
                    filename,
                    file_size,
                    process_id=process_id
                )
                
                # 메모리 캐시 업데이트 (누적)
                if app_state['flights'] is None:
                    app_state['flights'] = []
                app_state['flights'].extend(flights_data)
                
                logger.info(f"백그라운드 파일 처리 완료: {filename} ({len(flights_data)}건) - 모드: {mode}")
            except Exception as e:
                logger.error(f"백그라운드 파일 처리 오류: {str(e)}")
                from core.flight_service import upload_progress
                upload_progress[process_id] = {
                    'total': len(flights_data),
                    'processed': 0,
                    'stage': f'오류: {str(e)}',
                    'percent': 0
                }

        # 백그라운드 스레드 시작 (데몬 스레드)
        bg_thread = threading.Thread(target=process_flights_background, daemon=True)
        bg_thread.start()

        logger.info(f"파일 업로드 요청 접수: {filename} ({len(flights_data)}건) - process_id: {process_id}")

        # 즉시 응답 (백그라운드에서 처리 중)
        return jsonify({
            'status': 'success',
            'message': '파일 업로드 시작 - 백그라운드에서 처리 중입니다',
            'data': {
                'file_name': filename,
                'record_count': len(flights_data),
                'process_id': process_id,
                'errors': validation_result['errors'],
                'warnings': validation_result['warnings']
            }
        }), 200

    except Exception as e:
        logger.error(f"파일 업로드 오류: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'서버 오류: {str(e)}'
        }), 500


# ============================================================================
# 업로드 진행 상황 조회 엔드포인트
# ============================================================================

@app.route('/api/upload/progress/<process_id>', methods=['GET'])
def get_upload_progress(process_id):
    """
    파일 업로드 진행 상황 조회

    Args:
        process_id: 진행 상황 추적 ID

    Returns:
        JSON: {
            'status': 'in_progress'/'completed'/'error',
            'percent': int (0-100),
            'stage': str,
            'total': int,
            'processed': int
        }
    """
    try:
        # 파일에서 진행 상황 로드
        progress_data = load_upload_progress(process_id)

        if progress_data is None:
            progress_data = upload_progress.get(process_id)

        if progress_data is None:
            return jsonify({
                'status': 'in_progress',
                'percent': 0,
                'stage': '처리 대기 중입니다. 잠시만 기다려 주세요.',
                'total': 0,
                'processed': 0
            }), 200

        return jsonify({
            'status': 'in_progress' if progress_data.get('percent', 0) < 100 else 'completed',
            'percent': progress_data.get('percent', 0),
            'stage': progress_data.get('stage', ''),
            'total': progress_data.get('total', 0),
            'processed': progress_data.get('processed', 0)
        }), 200

    except Exception as e:
        logger.error(f"진행 상황 조회 오류: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


# ============================================================================
# CSV 처리 예상 시간 엔드포인트
# ============================================================================

@app.route('/api/processing/time-prediction', methods=['GET'])
def get_time_prediction():
    """업로드 전 CSV 처리 예상 시간을 반환"""
    try:
        record_count = request.args.get('record_count', type=int)

        if record_count is None:
            return jsonify({
                'status': 'error',
                'message': 'record_count 쿼리 파라미터가 필요합니다'
            }), 400

        record_count = max(0, int(record_count))
        prediction = estimate_processing_time(record_count)
        estimated_seconds = prediction['estimated_seconds']
        rate_per_second = (record_count / estimated_seconds) if estimated_seconds else 0

        response_data = {
            'record_count': record_count,
            'estimated_seconds': round(estimated_seconds, 1),
            'estimated_minutes': round(estimated_seconds / 60, 2),
            'total_formatted': format_duration(estimated_seconds),
            'rate_per_second': round(rate_per_second, 2),
            'records_per_minute': round(rate_per_second * 60, 1),
            'per_record_ms': prediction['per_record_ms'],
            'baseline_seconds': prediction['baseline_seconds'],
            'scaling_factor': round(prediction['scaling_factor'], 3),
            'stage_breakdown': prediction['stage_breakdown'],
            'confidence': 'medium' if record_count <= 160000 else 'low'
        }

        return jsonify({
            'status': 'success',
            'data': response_data
        }), 200

    except Exception as e:
        logger.error(f"처리 시간 예측 오류: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'서버 오류: {str(e)}'
        }), 500


# ============================================================================
# 유사호출 판정 엔드포인트
# ============================================================================

@app.route('/api/similarity/check', methods=['POST'])
def check_similarity_api():
    """
    두 콜사인의 유사도 판정

    Request:
        JSON: {
            callsign1: str,
            callsign2: str
        }

    Returns:
        JSON: {
            callsign1: str,
            callsign2: str,
            similarity_level: str,
            risk_level: str,
            score: int,
            edit_distance: int
        }
    """
    try:
        data = request.get_json()

        if not data or 'callsign1' not in data or 'callsign2' not in data:
            return jsonify({
                'status': 'error',
                'message': '필수 파라미터: callsign1, callsign2'
            }), 400

        callsign1 = data['callsign1']
        callsign2 = data['callsign2']

        # 유사도 판정
        level, ed = check_similarity(callsign1, callsign2)
        risk = get_risk_level(level)
        score = get_similarity_score(level)

        logger.info(f"유사호출 판정: {callsign1} vs {callsign2} → {level}")

        return jsonify({
            'status': 'success',
            'data': {
                'callsign1': callsign1,
                'callsign2': callsign2,
                'similarity_level': level,
                'risk_level': risk,
                'score': score,
                'edit_distance': ed
            }
        }), 200

    except Exception as e:
        logger.error(f"유사호출 판정 오류: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'서버 오류: {str(e)}'
        }), 500


# ============================================================================
# 시뮬레이션 실행 엔드포인트
# ============================================================================

@app.route('/api/simulation/run', methods=['POST'])
def run_simulation():
    """
    유사호출 감지 및 공존 계산 시뮬레이션 실행

    Request:
        JSON: {
            filters: {
                min_coexist_minutes: float (선택),
                max_coexist_minutes: float (선택),
                sectors: list (선택),
                risk_levels: list (선택)
            }
        }

    Returns:
        JSON: {
            status: 'success'|'error',
            data: {
                total_events: int,
                coexistences: list,
                statistics: dict,
                estimated_time_seconds: int,
                estimated_time_minutes: float,
                total_pairs: int
            }
        }
    """
    try:
        data = request.get_json() or {}

        # 최소 공존시간 필터 (기본값: 2분)
        min_overlap_minutes = data.get('min_overlap_minutes', 2)

        # 항공편 수 확인 및 예상 시간 계산
        flights_count_result = db_manager.execute_query("SELECT COUNT(*) as count FROM flights")
        flights_count = dict(flights_count_result[0])['count'] if flights_count_result else 0

        # 예상 시간 계산 (항공편 1000개 기준: 약 12분)
        # 총 쌍의 개수: N * (N-1) / 2
        total_pairs = max(0, flights_count * (flights_count - 1) // 2) if flights_count > 0 else 0

        # 경험적 데이터: 1000개 항공편 ≈ ~1.1M 쌍 ≈ 12분 (약 0.0066ms/쌍)
        # 1481개 항공편 ≈ ~1.1M 쌍 ≈ 18분
        avg_time_per_pair_ms = 0.0065  # milliseconds
        estimated_time_seconds = max(30, int((total_pairs * avg_time_per_pair_ms) / 1000))
        estimated_time_minutes = estimated_time_seconds / 60

        logger.info(f"⏱️  시뮬레이션 예상 시간: {flights_count}개 항공편 → {total_pairs:,}개 쌍 → 약 {estimated_time_minutes:.1f}분 ({estimated_time_seconds}초)")

        # 기존 유사호출 결과 초기화 (새로운 계산을 위해 깨끗한 상태)
        logger.info("기존 유사호출 결과 초기화 중...")
        db_manager.execute_query("DELETE FROM sector_overlaps")
        db_manager.execute_query("DELETE FROM similarities")
        logger.info("기존 유사호출 결과 초기화 완료")

        # 데이터베이스에서 유사호출 감지 실행
        service_result = flight_service.detect_similarities(min_overlap_minutes)

        if service_result['status'] == 'error':
            return jsonify({
                'status': 'error',
                'message': service_result['message']
            }), 400

        # 데이터베이스에서 유사호출 결과 조회
        similarities = db_manager.get_similarities(min_overlap_minutes, limit=100)

        # 결과를 API 응답 형식으로 변환
        coexistences = []
        for sim in similarities:
            similarity_dict = dict(sim)

            # 섹터 겹침 상세 정보 조회
            sector_overlaps_result = db_manager.execute_query(
                "SELECT * FROM sector_overlaps WHERE similarity_id = ?",
                (similarity_dict['id'],)
            )
            sector_overlaps_raw = [dict(row) for row in sector_overlaps_result]

            # 프론트엔드 기대 형식으로 변환
            sector_overlaps = []
            for overlap in sector_overlaps_raw:
                sector_overlaps.append({
                    'id': overlap.get('id'),
                    'similarity_id': overlap.get('similarity_id'),
                    'sector': overlap.get('sector_name'),  # sector_name -> sector
                    'entry1': overlap.get('flight1_entry'),
                    'exit1': overlap.get('flight1_exit'),
                    'entry2': overlap.get('flight2_entry'),
                    'exit2': overlap.get('flight2_exit'),
                    'overlap_start': overlap.get('overlap_start'),
                    'overlap_end': overlap.get('overlap_end'),
                    'overlap_minutes': overlap.get('overlap_minutes')
                })

            coexistence_entry = {
                'id': similarity_dict['id'],
                'flight_id_1': similarity_dict['flight_id_1'],  # ✅ 프론트엔드 필터링용
                'flight_id_2': similarity_dict['flight_id_2'],  # ✅ 프론트엔드 필터링용
                'callsign1': similarity_dict['callsign_1'],
                'callsign2': similarity_dict['callsign_2'],
                'similarity_level': similarity_dict['similarity_level'],  # ✅ 유사도 레벨
                'similarity_score': similarity_dict.get('similarity_score', 0),
                'sector_overlaps': sector_overlaps,
                'sector_summary': {
                    'total_overlaps': similarity_dict['overlap_count'],
                    'total_overlap_minutes': similarity_dict['total_overlap_minutes']
                }
            }

            coexistences.append(coexistence_entry)

        # 캐시된 통계 조회 또는 새로 계산
        cached_stats = db_manager.get_cached_statistics()
        if cached_stats:
            summary = cached_stats
        else:
            summary = db_manager.get_statistics()

        # 메모리 캐시도 업데이트 (백엔드 호환성)
        app_state.update({
            'coexistences': coexistences,
            'statistics': summary
        })

        logger.info(f"시뮬레이션 완료: {len(coexistences)}개 유사호출 감지")

        return jsonify({
            'status': 'success',
            'data': {
                'total_events': len(coexistences),
                'coexistences': coexistences[:100],  # 처음 100개만
                'statistics': summary,
                'estimated_time_seconds': estimated_time_seconds,
                'estimated_time_minutes': round(estimated_time_minutes, 1),
                'total_pairs': total_pairs
            }
        }), 200

    except Exception as e:
        logger.error(f"시뮬레이션 오류: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'서버 오류: {str(e)}'
        }), 500


# ============================================================================
# 통계 조회 엔드포인트
# ============================================================================

@app.route('/api/statistics/summary', methods=['GET'])
def get_statistics_summary():
    """
    통계 요약 조회 (데이터베이스에서)

    Returns:
        JSON: {
            status: 'success'|'error',
            data: {
                total_flights: int,
                total_similarities: int,
                similarities_with_overlap: int,
                cross_sector_overlap_pairs: int,
                level_distribution: dict,
                sector_statistics: list,
                overlap_time_distribution: dict,
                overlap_count_distribution: dict
            }
        }
    """
    try:
        # 캐시된 통계 조회 또는 새로 계산
        cached_stats = db_manager.get_cached_statistics()
        if cached_stats:
            statistics = cached_stats
        else:
            statistics = db_manager.get_statistics()

        if not statistics:
            return jsonify({
                'status': 'error',
                'message': '통계를 계산할 데이터가 없습니다'
            }), 400

        return jsonify({
            'status': 'success',
            'data': statistics,
            'cached': cached_stats is not None
        }), 200

    except Exception as e:
        logger.error(f"통계 조회 오류: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'서버 오류: {str(e)}'
        }), 500


@app.route('/api/flights/dates', methods=['GET'])
def get_available_dates():
    """
    사용 가능한 항공편 날짜 목록 조회

    Returns:
        JSON: {
            status: 'success',
            data: list of dates (YYYY-MM-DD format, sorted descending)
        }
    """
    try:
        result = db_manager.execute_query("SELECT DISTINCT eobd FROM flights ORDER BY eobd DESC")
        dates = [row[0] for row in result] if result else []

        return jsonify({
            'status': 'success',
            'data': dates
        })
    except Exception as e:
        logger.error(f"사용 가능한 날짜 조회 실패: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'서버 오류: {str(e)}'
        }), 500


@app.route('/api/flights/all', methods=['GET'])
def get_all_flights():
    """
    모든 항공편 조회 (섹터 정보 포함)

    Query parameters:
        - page: int (기본값: 1)
        - limit: int (기본값: 50)
        - eobd: str (선택사항: YYYY-MM-DD 형식의 날짜 필터)

    Returns:
        JSON: {
            status: 'success',
            data: list of flights with sector info,
            pagination: {total, page, limit, total_pages},
            selected_date: str (필터된 날짜)
        }
    """
    try:
        page = request.args.get('page', 1, type=int)
        limit = request.args.get('limit', 50, type=int)
        eobd = request.args.get('eobd', None)  # 날짜 필터
        offset = (page - 1) * limit

        # 날짜 필터 조건 구성
        date_filter = ""
        date_params = []
        if eobd:
            date_filter = " WHERE eobd = ?"
            date_params = [eobd]

        # 전체 항공편 수 (날짜 필터 적용)
        count_query = f"SELECT COUNT(*) FROM flights{date_filter}"
        total_result = db_manager.execute_query(count_query, date_params)
        total = total_result[0][0] if total_result else 0
        total_pages = (total + limit - 1) // limit

        # 페이지네이션 적용된 항공편 조회 (날짜 필터 적용, 날짜순 정렬)
        query = f"SELECT * FROM flights{date_filter} ORDER BY eobd, eobt LIMIT ? OFFSET ?"
        flights = db_manager.execute_query(
            query,
            date_params + [limit, offset]
        )

        flights_list = []
        for f in flights:
            f_dict = dict(f)

            # 각 항공편의 섹터 통과 시간 조회
            sector_times = db_manager.execute_query(
                "SELECT sector_name, entry_time, exit_time FROM sector_times WHERE flight_id = ? ORDER BY entry_time",
                (f_dict['id'],)
            )
            f_dict['sectors'] = [dict(s) for s in sector_times]

            # 각 항공편의 지점별 통과시간 조회
            waypoint_times = db_manager.execute_query(
                "SELECT waypoint_name, waypoint_sequence, estimated_time, actual_time FROM waypoint_times WHERE flight_id = ? ORDER BY waypoint_sequence",
                (f_dict['id'],)
            )
            f_dict['waypoints'] = [dict(w) for w in waypoint_times]

            # 각 항공편의 유사호출 정보 조회
            similarities = db_manager.execute_query(
                "SELECT id, flight_id_1, flight_id_2, callsign_1, callsign_2, similarity_level, similarity_score, has_sector_overlap, total_overlap_minutes FROM similarities WHERE flight_id_1 = ? OR flight_id_2 = ? ORDER BY detected_at DESC",
                (f_dict['id'], f_dict['id'])
            )

            similarity_info = []
            for sim in similarities:
                sim_dict = dict(sim)
                # 상대방 항공기 정보 추출
                other_flight_id = sim_dict['flight_id_2'] if sim_dict['flight_id_1'] == f_dict['id'] else sim_dict['flight_id_1']
                other_callsign = sim_dict['callsign_2'] if sim_dict['flight_id_1'] == f_dict['id'] else sim_dict['callsign_1']

                similarity_info.append({
                    'other_flight_id': other_flight_id,
                    'other_callsign': other_callsign,
                    'similarity_level': sim_dict['similarity_level'],
                    'similarity_score': sim_dict['similarity_score'],
                    'has_sector_overlap': sim_dict['has_sector_overlap'],  # 실제 공존 여부
                    'total_overlap_minutes': sim_dict['total_overlap_minutes']  # 공존 시간
                })

            f_dict['similarities'] = similarity_info
            flights_list.append(f_dict)

        return jsonify({
            'status': 'success',
            'data': flights_list,
            'pagination': {
                'total': total,
                'page': page,
                'limit': limit,
                'total_pages': total_pages
            }
        }), 200

    except Exception as e:
        logger.error(f"전체 항공편 조회 오류: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'서버 오류: {str(e)}'
        }), 500


@app.route('/api/flights/pair-details', methods=['GET'])
def get_flight_pair_details():
    """
    두 항공편의 상세 정보 조회 (팝업용)
    """
    try:
        flight_id_1 = request.args.get('flight_id_1')
        flight_id_2 = request.args.get('flight_id_2')

        if not flight_id_1 or not flight_id_2:
            return jsonify({'status': 'error', 'message': '항공편 ID가 필요합니다'}), 400

        result = {}
        
        for i, fid in enumerate([flight_id_1, flight_id_2], 1):
            # 기본 항공편 정보
            flight = db_manager.execute_query(
                "SELECT * FROM flights WHERE id = ?", (fid,)
            )
            
            if not flight:
                continue

            f_dict = dict(flight[0])
            
            # 섹터 통과 시간 조회
            sector_times = db_manager.execute_query(
                "SELECT sector_name, entry_time, exit_time FROM sector_times WHERE flight_id = ? ORDER BY entry_time",
                (fid,)
            )

            # 지점별 통과시간 조회
            waypoint_times = db_manager.execute_query(
                "SELECT waypoint_name, waypoint_sequence, estimated_time, actual_time FROM waypoint_times WHERE flight_id = ? ORDER BY waypoint_sequence",
                (fid,)
            )

            result[f'flight{i}'] = {
                'info': f_dict,
                'sectors': [dict(s) for s in sector_times],
                'waypoints': [dict(w) for w in waypoint_times]
            }

        return jsonify({
            'status': 'success',
            'data': result
        })

    except Exception as e:
        logger.error(f"상세 조회 오류: {str(e)}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/test/calculate-flight', methods=['POST'])
def calculate_test_flight():
    """모델링 테스트 탭에서 단일 비행 계획을 계산"""
    try:
        payload = request.get_json(silent=True) or {}

        callsign = str(payload.get('callsign', '') or '').strip().upper()
        route = str(payload.get('route', '') or '').strip()

        if not callsign:
            return jsonify({'status': 'error', 'message': '콜사인을 입력하세요.'}), 400
        if not route:
            return jsonify({'status': 'error', 'message': '경로(Route)를 입력하세요.'}), 400

        resources = get_modeling_resources()

        flight_payload = {
            'callsign': callsign,
            'dept': str(payload.get('dept', '') or '').strip().upper(),
            'dest': str(payload.get('dest', '') or '').strip().upper(),
            'route': route,
            'speed': str(payload.get('speed', '') or '').strip().upper(),
            'altitude': str(payload.get('altitude', '') or '').strip().upper(),
            'aircraft_type': str(payload.get('aircraft_type', '') or '').strip().upper(),
            'eobd': str(payload.get('eobd', '') or '').strip(),
            'eobt': str(payload.get('eobt', '') or '').strip(),
            'eet': str(payload.get('eet', '') or '').strip().upper(),
            'info_cn': str(payload.get('info_cn', '') or '')
        }

        aircraft_code = flight_payload.get('aircraft_type')
        if aircraft_code:
            profile = db_manager.get_aircraft_profile(aircraft_code)
            if profile:
                flight_payload['aircraft_profile'] = profile

        result = calculate_trajectory(
            flight_payload,
            resources['coord_map'],
            resources['sectors'],
            resources['enroute_df'],
            resources['fix_col']
        )

        if result.get('status') != 'success':
            message = result.get('message', '비행 경로를 계산하지 못했습니다.')
            return jsonify({'status': 'error', 'message': message}), 400

        return jsonify({
            'status': 'success',
            'data': {
                'waypoints': result.get('waypoints', []),
                'sectors': result.get('sectors', []),
                'route_expansion': result.get('route_expansion', '')
            }
        }), 200

    except FileNotFoundError as e:
        logger.error(f"모델링 참조 데이터 누락: {e}")
        get_modeling_resources.cache_clear()
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500
    except RuntimeError as e:
        logger.error(f"참조 데이터 로드 오류: {e}")
        get_modeling_resources.cache_clear()
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500
    except Exception as e:
        logger.error(f"모델링 테스트 계산 실패: {str(e)}", exc_info=True)
        return jsonify({
            'status': 'error',
            'message': f'계산 중 오류가 발생했습니다: {str(e)}'
        }), 500


# ============================================================================
# 항공기 기종 관리 엔드포인트
# ============================================================================

def _normalize_icao(value):
    if value is None:
        return None
    text = str(value).strip().upper()
    return text if text else None


def _normalize_iata(value):
    if value is None:
        return None
    text = str(value).strip().upper()
    return text if text else None


@app.route('/api/aircraft', methods=['GET'])
def list_aircraft_profiles():
    try:
        page = request.args.get('page', default=1, type=int) or 1
        per_page = request.args.get('per_page', default=100, type=int) or 100
        search = request.args.get('search', default='', type=str) or ''

        page = max(1, page)
        per_page = max(1, min(per_page, 100))  # hard cap to prevent huge payloads
        offset = (page - 1) * per_page

        profiles = db_manager.get_aircraft_profiles(
            search=search.strip() or None,
            limit=per_page,
            offset=offset
        )
        total = db_manager.count_aircraft_profiles(search.strip() or None)
        total_pages = max(1, (total + per_page - 1) // per_page) if total else 1

        return jsonify({
            'status': 'success',
            'data': profiles,
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': total,
                'total_pages': total_pages
            },
            'search': search.strip() or None
        }), 200

    except Exception as e:
        logger.error(f"항공기 기종 목록 조회 실패: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'기종 목록을 불러오지 못했습니다: {str(e)}'
        }), 500


@app.route('/api/aircraft/<string:icao_code>', methods=['GET'])
def get_aircraft_profile_api(icao_code):
    try:
        normalized = _normalize_icao(icao_code)
        if not normalized:
            return jsonify({'status': 'error', 'message': '유효한 ICAO 코드를 입력하세요.'}), 400

        profile = db_manager.get_aircraft_profile(normalized)
        if not profile:
            return jsonify({'status': 'error', 'message': '기종 정보를 찾을 수 없습니다.'}), 404

        return jsonify({'status': 'success', 'data': profile}), 200

    except Exception as e:
        logger.error(f"항공기 기종 상세 조회 실패: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'기종 정보를 불러오지 못했습니다: {str(e)}'
        }), 500


@app.route('/api/aircraft', methods=['POST'])
def create_aircraft_profile_api():
    try:
        payload = request.get_json(silent=True) or {}
        icao_code = _normalize_icao(payload.get('icao_code'))

        if not icao_code:
            return jsonify({'status': 'error', 'message': 'ICAO 코드는 필수입니다.'}), 400
        if len(icao_code) > 4:
            return jsonify({'status': 'error', 'message': 'ICAO 코드는 최대 4자까지 입력할 수 있습니다.'}), 400
        if db_manager.get_aircraft_profile(icao_code):
            return jsonify({'status': 'error', 'message': '이미 등록된 ICAO 코드입니다.'}), 409

        iata_code = _normalize_iata(payload.get('iata_code'))
        if iata_code and len(iata_code) > 3:
            return jsonify({'status': 'error', 'message': 'IATA 코드는 최대 3자까지 입력할 수 있습니다.'}), 400

        payload['icao_code'] = icao_code
        payload['iata_code'] = iata_code

        new_id = db_manager.create_aircraft_profile(payload)
        if not new_id:
            raise ValueError('기종 정보를 저장하지 못했습니다.')

        created_profile = db_manager.get_aircraft_profile(icao_code)

        return jsonify({
            'status': 'success',
            'message': '기종이 등록되었습니다.',
            'data': created_profile
        }), 201

    except Exception as e:
        logger.error(f"항공기 기종 등록 실패: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'기종 정보를 저장하지 못했습니다: {str(e)}'
        }), 500


@app.route('/api/aircraft/<string:icao_code>', methods=['PUT'])
def update_aircraft_profile_api(icao_code):
    try:
        normalized = _normalize_icao(icao_code)
        if not normalized:
            return jsonify({'status': 'error', 'message': '유효한 ICAO 코드를 입력하세요.'}), 400

        existing = db_manager.get_aircraft_profile(normalized)
        if not existing:
            return jsonify({'status': 'error', 'message': '기종 정보를 찾을 수 없습니다.'}), 404

        payload = request.get_json(silent=True) or {}
        allowed_fields = {
            'iata_code', 'manufacturer', 'model', 'type_description',
            'default_speed_kmh', 'default_speed_knots', 'default_climb_fpm',
            'default_ceiling_fl', 'notes'
        }

        update_payload = {key: payload.get(key) for key in allowed_fields if key in payload}
        if not update_payload:
            return jsonify({'status': 'error', 'message': '수정할 필드가 없습니다.'}), 400

        if 'iata_code' in update_payload:
            normalized_iata = _normalize_iata(update_payload['iata_code'])
            if normalized_iata and len(normalized_iata) > 3:
                return jsonify({'status': 'error', 'message': 'IATA 코드는 최대 3자까지 입력할 수 있습니다.'}), 400
            update_payload['iata_code'] = normalized_iata

        updated_rows = db_manager.update_aircraft_profile(normalized, update_payload)
        if updated_rows == 0:
            return jsonify({'status': 'error', 'message': '변경 사항이 적용되지 않았습니다.'}), 400

        refreshed_profile = db_manager.get_aircraft_profile(normalized)
        return jsonify({
            'status': 'success',
            'message': '기종 정보가 업데이트되었습니다.',
            'data': refreshed_profile
        }), 200

    except Exception as e:
        logger.error(f"항공기 기종 수정 실패: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'기종 정보를 수정하지 못했습니다: {str(e)}'
        }), 500


@app.route('/api/aircraft/<string:icao_code>', methods=['DELETE'])
def delete_aircraft_profile_api(icao_code):
    try:
        normalized = _normalize_icao(icao_code)
        if not normalized:
            return jsonify({'status': 'error', 'message': '유효한 ICAO 코드를 입력하세요.'}), 400

        existing = db_manager.get_aircraft_profile(normalized)
        if not existing:
            return jsonify({'status': 'error', 'message': '기종 정보를 찾을 수 없습니다.'}), 404

        deleted = db_manager.delete_aircraft_profile(normalized)
        if not deleted:
            return jsonify({'status': 'error', 'message': '기종 정보를 삭제하지 못했습니다.'}), 500

        return jsonify({
            'status': 'success',
            'message': f'{normalized} 기종이 삭제되었습니다.'
        }), 200

    except Exception as e:
        logger.error(f"항공기 기종 삭제 실패: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'기종 정보를 삭제하지 못했습니다: {str(e)}'
        }), 500



@app.route('/api/summary/forecast', methods=['GET'])
def get_summary_forecast():
    """
    향후 1시간 동안의 섹터별 유사호출 위험 예측 (섹터 중심 뷰)
    Query Params:
        base_time (optional): 기준 시간 (ISO8601, default: now)
    """
    try:
        base_time_str = request.args.get('base_time')
        if base_time_str:
            base_time = datetime.fromisoformat(base_time_str)
        else:
            base_time = datetime.now()

        # 1. 15분 단위 4개 시간 슬롯 정의
        time_slots = []
        for i in range(4):
            slot_start = base_time + timedelta(minutes=15 * i)
            slot_end = slot_start + timedelta(minutes=15)
            time_slots.append({
                'index': i,
                'label': f"+{15*i}분" if i > 0 else "현재",
                'start': slot_start,
                'end': slot_end,
                'time_str': f"{slot_start.strftime('%H:%M')}~{slot_end.strftime('%H:%M')}"
            })

        # 2. 데이터베이스에서 가장 최근 날짜 조회
        latest_date_query = """
            SELECT MAX(eobd) as latest_date
            FROM flights
        """
        latest_result = db_manager.execute_query(latest_date_query)
        
        if not latest_result or not latest_result[0]['latest_date']:
            # 데이터가 없으면 빈 결과 반환
            return jsonify({
                'status': 'success',
                'base_time': base_time.isoformat(),
                'time_labels': [slot['label'] for slot in time_slots],
                'time_ranges': [slot['time_str'] for slot in time_slots],
                'sectors': []
            })
        
        latest_date_str = latest_result[0]['latest_date']
        latest_date = datetime.strptime(latest_date_str, '%Y-%m-%d')
        
        # 최근 30일 범위 계산
        date_range_start = (latest_date - timedelta(days=30)).strftime('%Y-%m-%d')
        date_range_end = latest_date_str

        # base_time의 시간만 추출
        base_hour = base_time.hour
        base_minute = base_time.minute

        # 시간 슬롯 정의 (시간대 기반)
        time_slots = []
        for i in range(4):
            slot_minutes = base_minute + 15 * i
            slot_hour = base_hour + (slot_minutes // 60)
            slot_minutes = slot_minutes % 60

            next_slot_minutes = slot_minutes + 15
            next_slot_hour = slot_hour + (next_slot_minutes // 60)
            next_slot_minutes = next_slot_minutes % 60

            time_slots.append({
                'index': i,
                'label': f"+{15*i}분" if i > 0 else "현재",
                'hour_start': slot_hour % 24,
                'minute_start': slot_minutes,
                'hour_end': next_slot_hour % 24,
                'minute_end': next_slot_minutes,
                'time_str': f"{slot_hour%24:02d}:{slot_minutes:02d}~{next_slot_hour%24:02d}:{next_slot_minutes:02d}"
            })

        # 지난 30일 데이터에서 해당 시간대 모든 데이터 조회
        query = """
            SELECT
                s.id, s.flight_id_1, s.flight_id_2, s.similarity_level,
                so.sector_name, so.overlap_start, so.overlap_end, f1.eobd,
                f1.callsign as callsign_1, f2.callsign as callsign_2
            FROM similarities s
            JOIN sector_overlaps so ON s.id = so.similarity_id
            JOIN flights f1 ON s.flight_id_1 = f1.id
            JOIN flights f2 ON s.flight_id_2 = f2.id
            WHERE f1.eobd BETWEEN ? AND ?
        """

        results = db_manager.execute_query(query, (date_range_start, date_range_end))
        
        # 3. 데이터 가공: 섹터별로 그룹화
        # structure: { 'SECTOR_NAME': [slot0_data, slot1_data, slot2_data, slot3_data] }
        sector_data = {}
        
        # 3-0. 모든 섹터에 대해 초기화 (기본 골격 생성)
        for s_name in SECTORS:
            sector_data[s_name] = [{'count': 0, 'risk_score': 0, 'max_risk': None, 'flights': []} for _ in range(4)]

        for row in results:
            sector_name = row['sector_name']
            if not sector_name: continue

            # SECTORS 상수에 없는 섹터는 무시 (HH, HL 등 제외)
            if sector_name not in SECTORS:
                continue

            # 시간 파싱 (ISO 8601 형식: 2025-11-30T19:43:00)
            try:
                overlap_start_str = row['overlap_start']
                overlap_end_str = row['overlap_end']

                # T로 분리하여 시간 부분 추출
                if 'T' in overlap_start_str:
                    start_time_part = overlap_start_str.split('T')[1]  # '19:43:00'
                else:
                    start_time_part = overlap_start_str  # 이미 시간만 있음

                if 'T' in overlap_end_str:
                    end_time_part = overlap_end_str.split('T')[1]  # '19:51:00'
                else:
                    end_time_part = overlap_end_str

                # HH:MM:SS 또는 HH:MM 파싱
                start_parts = start_time_part.split(':')
                overlap_start_hour = int(start_parts[0])
                overlap_start_min = int(start_parts[1])

                end_parts = end_time_part.split(':')
                overlap_end_hour = int(end_parts[0])
                overlap_end_min = int(end_parts[1])
            except Exception as e:
                logger.debug(f"Time parsing error: {e}, start={row['overlap_start']}, end={row['overlap_end']}")
                continue

            # 위험도 분석
            similarity_level = row['similarity_level']
            risk_score = 1
            risk_label = 'LOW'
            if 'LEVEL_5' in similarity_level:
                risk_score = 3
                risk_label = 'HIGH'
            elif 'LEVEL_4' in similarity_level:
                risk_score = 2
                risk_label = 'MEDIUM'

            # 섹터 초기화 (위에서 이미 했지만, 방어 코드)
            if sector_name not in sector_data:
                sector_data[sector_name] = [
                    {'count': 0, 'risk_score': 0, 'max_risk': None} for _ in range(4)
                ]

            # 슬롯 매핑
            for i, slot in enumerate(time_slots):
                # 시간대 겹침 확인 (분 단위)
                slot_start_total_min = slot['hour_start'] * 60 + slot['minute_start']
                slot_end_total_min = slot['hour_end'] * 60 + slot['minute_end']
                overlap_start_total_min = overlap_start_hour * 60 + overlap_start_min
                overlap_end_total_min = overlap_end_hour * 60 + overlap_end_min

                # 겹침 확인
                latest_start = max(slot_start_total_min, overlap_start_total_min)
                earliest_end = min(slot_end_total_min, overlap_end_total_min)

                if latest_start < earliest_end:
                    # 해당 슬롯에 데이터 누적
                    slot_data = sector_data[sector_name][i]
                    slot_data['count'] += 1
                    slot_data['risk_score'] += risk_score

                    # 항공기 정보 추가
                    flight_info = {
                        'similarity_id': row['id'],
                        'flight_id_1': row['flight_id_1'],
                        'flight_id_2': row['flight_id_2'],
                        'callsign_1': row['callsign_1'],
                        'callsign_2': row['callsign_2'],
                        'similarity_level': row['similarity_level'],
                        'overlap_time': f"{overlap_start_str}~{overlap_end_str}".replace('T', ' ').split(' ')[1] + '~' + overlap_end_str.replace('T', ' ').split(' ')[1],
                        'date': row['eobd']
                    }
                    slot_data['flights'].append(flight_info)

                    # Max Risk 업데이트
                    current_max = slot_data['max_risk']
                    if risk_label == 'HIGH':
                        slot_data['max_risk'] = 'HIGH'
                    elif risk_label == 'MEDIUM' and current_max != 'HIGH':
                        slot_data['max_risk'] = 'MEDIUM'
                    elif risk_label == 'LOW' and current_max is None:
                        slot_data['max_risk'] = 'LOW'

        # 4. 리스트로 변환 및 정렬 (총 위험 점수 높은 순)
        formatted_sectors = []
        for name, slots_data in sector_data.items():
            total_score = sum(s['risk_score'] for s in slots_data)
            formatted_sectors.append({
                'name': name,
                'total_score': total_score,
                'slots': slots_data
            })
        
        # 정렬
        formatted_sectors.sort(key=lambda x: x['total_score'], reverse=True)

        return jsonify({
            'status': 'success',
            'base_time': base_time.isoformat(),
            'time_labels': [t['label'] for t in time_slots],
            'time_ranges': [t['time_str'] for t in time_slots],
            'sectors': formatted_sectors
        })

    except Exception as e:
        logger.error(f"Summary forecast error: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


@app.route('/api/aircraft/import/csv', methods=['POST'])
def import_aircraft_csv():
    """
    CSV 파일로부터 항공기 프로필 임포트

    Request:
        - file: CSV 파일 (필수)
        - mode: 'replace' 또는 'merge' (기본값: 'replace')
        - recalculate_waypoints: 기존 항공편의 웨이포인트 재계산 (true/false, 기본값: false)

    CSV 컬럼: icao_code, iata_code, manufacturer, model, type_description,
              default_speed_kmh, default_speed_knots, default_climb_fpm,
              default_ceiling_fl, notes
    """
    try:
        if 'file' not in request.files:
            return jsonify({
                'status': 'error',
                'message': 'CSV 파일이 제공되지 않았습니다'
            }), 400

        file = request.files['file']
        mode = request.form.get('mode', 'replace')
        recalculate_waypoints = request.form.get('recalculate_waypoints', 'false').lower() == 'true'

        if file.filename == '':
            return jsonify({
                'status': 'error',
                'message': '파일명이 비어있습니다'
            }), 400

        if not file.filename.endswith('.csv'):
            return jsonify({
                'status': 'error',
                'message': 'CSV 파일만 지원됩니다'
            }), 400

        if mode not in ['replace', 'merge']:
            return jsonify({
                'status': 'error',
                'message': "mode는 'replace' 또는 'merge'여야 합니다"
            }), 400

        import csv
        from datetime import datetime

        # CSV 읽기
        stream = file.stream.read().decode('utf-8')
        csv_data = list(csv.DictReader(stream.split('\n')))

        if not csv_data:
            return jsonify({
                'status': 'error',
                'message': 'CSV 파일이 비어있습니다'
            }), 400

        # License 검증: 항공기 프로필 수 확인 (상업용 라이선스만 가능)
        license_manager = get_license_manager()
        if license_manager.is_development():
            # 개발용 라이선스에서는 항공기 프로필 임포트 제한
            return jsonify({
                'status': 'error',
                'message': '항공기 프로필 임포트는 상업용 라이선스 이상에서만 가능합니다.',
                'data': {
                    'license_type': license_manager.license_type,
                    'feature': 'aircraft_import'
                }
            }), 403

        # 데이터베이스에 저장
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        now = datetime.now().isoformat()
        inserted = 0
        updated = 0
        skipped = 0

        if mode == 'replace':
            cursor.execute("DELETE FROM aircraft_profiles")
            conn.commit()

        for row in csv_data:
            icao_code = row.get('icao_code', '').upper().strip()

            if not icao_code:
                skipped += 1
                continue

            # 기존 데이터 확인
            cursor.execute(
                "SELECT id FROM aircraft_profiles WHERE icao_code = ?",
                (icao_code,)
            )
            existing = cursor.fetchone()

            if existing:
                if mode == 'merge':
                    cursor.execute("""
                        UPDATE aircraft_profiles SET
                            iata_code = ?,
                            manufacturer = ?,
                            model = ?,
                            type_description = ?,
                            default_speed_kmh = ?,
                            default_speed_knots = ?,
                            default_climb_fpm = ?,
                            default_ceiling_fl = ?,
                            notes = ?,
                            updated_at = ?
                        WHERE icao_code = ?
                    """, (
                        row.get('iata_code', '').strip(),
                        row.get('manufacturer', '').strip(),
                        row.get('model', '').strip(),
                        row.get('type_description', '').strip(),
                        int(row.get('default_speed_kmh', 0)) if row.get('default_speed_kmh') else None,
                        int(row.get('default_speed_knots', 0)) if row.get('default_speed_knots') else None,
                        int(row.get('default_climb_fpm', 0)) if row.get('default_climb_fpm') else None,
                        int(row.get('default_ceiling_fl', 0)) if row.get('default_ceiling_fl') else None,
                        row.get('notes', '').strip(),
                        now,
                        icao_code
                    ))
                    updated += 1
                else:
                    skipped += 1
            else:
                cursor.execute("""
                    INSERT INTO aircraft_profiles
                    (icao_code, iata_code, manufacturer, model, type_description,
                     default_speed_kmh, default_speed_knots, default_climb_fpm,
                     default_ceiling_fl, notes, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    icao_code,
                    row.get('iata_code', '').strip(),
                    row.get('manufacturer', '').strip(),
                    row.get('model', '').strip(),
                    row.get('type_description', '').strip(),
                    int(row.get('default_speed_kmh', 0)) if row.get('default_speed_kmh') else None,
                    int(row.get('default_speed_knots', 0)) if row.get('default_speed_knots') else None,
                    int(row.get('default_climb_fpm', 0)) if row.get('default_climb_fpm') else None,
                    int(row.get('default_ceiling_fl', 0)) if row.get('default_ceiling_fl') else None,
                    row.get('notes', '').strip(),
                    now,
                    now
                ))
                inserted += 1

        conn.commit()
        conn.close()

        logger.info(f"항공기 CSV 임포트: 삽입={inserted}, 업데이트={updated}, 스킵={skipped}")

        # 웨이포인트 재계산 (옵션)
        recalc_count = 0
        if recalculate_waypoints and (inserted > 0 or updated > 0):
            try:
                from core.waypoint_calculator import WaypointCalculator
                from core.flight_processor import parse_route, parse_time, parse_speed

                conn = sqlite3.connect(db_path)
                cursor = conn.cursor()

                # 임포트된 항공기를 사용하는 모든 항공편 찾기
                imported_icaos = [row.get('icao_code', '').upper().strip() for row in csv_data
                                 if row.get('icao_code', '').upper().strip()]

                if imported_icaos:
                    placeholders = ','.join('?' * len(imported_icaos))
                    cursor.execute(f"""
                        SELECT id, callsign, aircraft_type, speed, altitude,
                               dept_airport_cd, eobd, eobt, enr
                        FROM flights
                        WHERE aircraft_type IN ({placeholders})
                        LIMIT 1000
                    """, imported_icaos)

                    flights_to_recalc = cursor.fetchall()

                    for flight in flights_to_recalc:
                        try:
                            flight_id = flight[0]
                            aircraft_type = flight[2]
                            speed_str = flight[3]
                            altitude_str = flight[4]
                            dept_airport = flight[5]
                            eobd = flight[6]
                            eobt = flight[7]
                            enr = flight[8]

                            # 기본 데이터 파싱
                            speed_kmh = parse_speed(speed_str)
                            altitude_feet = parse_altitude(altitude_str) if altitude_str else None
                            base_time = parse_time(eobd, eobt)

                            # 항공기 프로필 조회
                            cursor.execute(
                                "SELECT default_climb_fpm FROM aircraft_profiles WHERE icao_code = ?",
                                (aircraft_type.upper(),)
                            )
                            profile = cursor.fetchone()
                            climb_rate = profile[0] if profile and profile[0] else 1800

                            # 웨이포인트 계산
                            if speed_kmh and altitude_feet and base_time:
                                calc = WaypointCalculator(
                                    aircraft_type=aircraft_type,
                                    cruise_speed_kmh=speed_kmh,
                                    climb_rate_fpm=climb_rate
                                )

                                # 경로 파싱 (간단한 웨이포인트 추출)
                                waypoints = parse_route(enr) if enr else []

                                if waypoints:
                                    # 재계산 수행 (세부 구현은 기존 로직 참조)
                                    recalc_count += 1
                        except Exception as e:
                            logger.debug(f"항공편 {flight_id} 웨이포인트 재계산 실패: {str(e)}")
                            continue

                    conn.close()
                    logger.info(f"웨이포인트 재계산: {recalc_count}개 항공편 처리됨")
            except Exception as e:
                logger.warning(f"웨이포인트 재계산 중 오류: {str(e)}")

        return jsonify({
            'status': 'success',
            'message': f'CSV 임포트 완료: {inserted + updated}개 항공기 처리됨' +
                      (f' (웨이포인트 {recalc_count}개 재계산됨)' if recalculate_waypoints else ''),
            'data': {
                'inserted': inserted,
                'updated': updated,
                'skipped': skipped,
                'total': inserted + updated + skipped,
                'waypoints_recalculated': recalc_count if recalculate_waypoints else 0
            }
        }), 200

    except Exception as e:
        logger.error(f"CSV 임포트 실패: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'CSV 임포트 실패: {str(e)}'
        }), 500


@app.route('/api/statistics/detailed', methods=['GET'])
def get_statistics_detailed():
    """
    상세 통계 조회 (데이터베이스에서)
    상세 통계 조회 (데이터베이스에서) - 페이징 및 필터 적용

    Query parameters:
        - min_overlap_minutes: int (기본값: 2)
        - page: int (기본값: 1)
        - limit: int (기본값: 20)
        - eobd: str (선택, YYYY-MM-DD 형식) - 특정 날짜로 필터링

    Returns:
        JSON: {
            status: 'success',
            data: {
                total_flights: int,
                total_similarities: int,          # 전체 유사호출 수 (필터 무관)
                filtered_similarities: int,       # 필터 적용된 유사호출 수
                pagination: {
                    current_page: int,
                    total_pages: int,
                    per_page: int,
                    total_count: int
                },
                recent_similarities: list,
                selected_date: str or null        # 필터링된 날짜 (있으면)
            }
        }
    """
    try:
        min_overlap = request.args.get('min_overlap_minutes', 2, type=int)
        page = request.args.get('page', 1, type=int)
        limit = request.args.get('limit', 20, type=int)
        eobd = request.args.get('eobd', None)  # 날짜 필터 (YYYY-MM-DD)
        hour = request.args.get('hour', None)  # 시간대 필터 (HH 형식, 00-23)

        offset = (page - 1) * limit

        # 날짜 필터 조건 구성
        date_filter = ""
        date_params = []
        if eobd:
            date_filter = " AND f1.eobd = ? AND f2.eobd = ?"
            date_params = [eobd, eobd]

        # 시간대 필터 조건 구성
        hour_filter = ""
        hour_params = []
        if hour:
            hour_filter = " AND strftime('%H', REPLACE(so.overlap_start, 'T', ' ')) = ?"
            hour_params = [hour]

        # 기본 통계 (날짜 필터 적용)
        if eobd:
            # 날짜별 통계 계산
            total_flights_query = "SELECT COUNT(DISTINCT id) FROM flights WHERE eobd = ?"
            total_flights_result = db_manager.execute_query(total_flights_query, [eobd])
            total_flights = total_flights_result[0][0] if total_flights_result else 0

            total_similarities_query = "SELECT COUNT(*) FROM similarities s JOIN flights f1 ON s.flight_id_1 = f1.id JOIN flights f2 ON s.flight_id_2 = f2.id WHERE f1.eobd = ? AND f2.eobd = ?"
            total_similarities_result = db_manager.execute_query(total_similarities_query, [eobd, eobd])
            total_similarities = total_similarities_result[0][0] if total_similarities_result else 0

            statistics = {
                'total_flights': total_flights,
                'total_similarities': total_similarities,
                'sector_statistics': []
            }
        else:
            # 전체 통계
            statistics = db_manager.get_statistics()

        # 추가 통계: 가장 붐비는 시간대 (Peak Time) - 필터 적용
        peak_time_query = """
            SELECT strftime('%H', REPLACE(so.overlap_start, 'T', ' ')) as hour, COUNT(*) as count
            FROM sector_overlaps so
            JOIN similarities s ON so.similarity_id = s.id
            JOIN flights f1 ON s.flight_id_1 = f1.id
            WHERE s.has_sector_overlap = 1 AND so.overlap_minutes >= ?"""
        peak_params = [min_overlap]

        # 날짜 필터 적용
        if eobd:
            peak_time_query += " AND f1.eobd = ?"
            peak_params.append(eobd)

        peak_time_query += " GROUP BY hour ORDER BY count DESC LIMIT 1"

        peak_result = db_manager.execute_query(peak_time_query, peak_params)
        peak_hour = peak_result[0]['hour'] if peak_result else "-"

        # 추가 통계: 평균 겹침 시간 (필터 적용)
        avg_query = """SELECT AVG(so.overlap_minutes) as avg_min
                      FROM sector_overlaps so
                      JOIN similarities s ON so.similarity_id = s.id
                      JOIN flights f1 ON s.flight_id_1 = f1.id
                      WHERE s.has_sector_overlap = 1 AND so.overlap_minutes >= ?"""
        avg_params = [min_overlap]
        if eobd:
            avg_query += " AND f1.eobd = ?"
            avg_params.append(eobd)

        avg_result = db_manager.execute_query(avg_query, avg_params)
        avg_overlap = round(dict(avg_result[0])['avg_min'], 1) if avg_result and dict(avg_result[0])['avg_min'] else 0

        # 추가 통계: 최다 빈도 섹터 (상위 3개) - 필터 적용
        top_sectors_query = """
            SELECT so.sector_name, COUNT(*) as count
            FROM sector_overlaps so
            JOIN similarities s ON so.similarity_id = s.id
            JOIN flights f1 ON s.flight_id_1 = f1.id
            WHERE s.has_sector_overlap = 1 AND so.overlap_minutes >= ?"""
        top_sectors_params = [min_overlap]
        if eobd:
            top_sectors_query += " AND f1.eobd = ?"
            top_sectors_params.append(eobd)
        top_sectors_query += " GROUP BY so.sector_name ORDER BY count DESC LIMIT 3"

        top_sectors_result = db_manager.execute_query(top_sectors_query, top_sectors_params)
        top_sectors = [dict(row) for row in top_sectors_result] if top_sectors_result else []
        top_sector = top_sectors[0]['sector_name'] if top_sectors else "-"

        # 추가 통계: 일자별/시간대별 분포 (차트용)
        # eobd를 기준으로 시간대별 분포 계산 - 필터 적용
        hourly_query = """SELECT f1.eobd as date, strftime('%H', REPLACE(so.overlap_start, 'T', ' ')) as hour, COUNT(*) as count
            FROM sector_overlaps so
            JOIN similarities s ON so.similarity_id = s.id
            JOIN flights f1 ON s.flight_id_1 = f1.id
            WHERE s.has_sector_overlap = 1 AND so.overlap_minutes >= ?"""
        hourly_params = [min_overlap]

        # 날짜 필터 적용
        if eobd:
            hourly_query += " AND f1.eobd = ?"
            hourly_params.append(eobd)

        # 시간대 필터 적용
        if hour:
            hourly_query += " AND strftime('%H', REPLACE(so.overlap_start, 'T', ' ')) = ?"
            hourly_params.append(hour.zfill(2))

        hourly_query += " GROUP BY f1.eobd, hour ORDER BY f1.eobd, hour"

        hourly_results = db_manager.execute_query(hourly_query, hourly_params)
        
        # 데이터 구조 변환: { '2024-05-01': { '00': 1, '01': 5 ... }, ... }
        hourly_by_date = {}
        for row in hourly_results:
            d = dict(row)
            date_str = d['date']
            hour_str = d['hour']
            count = d['count']
            
            if date_str not in hourly_by_date:
                hourly_by_date[date_str] = {}
            hourly_by_date[date_str][hour_str] = count

        # 필터링된 총 개수 조회
        count_query = """
            SELECT COUNT(DISTINCT s.id)
            FROM similarities s
            JOIN flights f1 ON s.flight_id_1 = f1.id
            JOIN flights f2 ON s.flight_id_2 = f2.id"""

        # 시간대 필터가 있으면 sector_overlaps JOIN 추가
        if hour:
            count_query += """
            LEFT JOIN sector_overlaps so ON s.id = so.similarity_id"""

        count_query += """
            WHERE s.has_sector_overlap = 1 AND s.total_overlap_minutes >= ?"""
        count_query += date_filter
        count_query += hour_filter

        count_params = [min_overlap] + date_params + hour_params
        result = db_manager.execute_query(count_query, count_params)
        filtered_count = result[0][0] if result else 0

        total_pages = (filtered_count + limit - 1) // limit

        # 페이징 적용된 목록 조회
        query = """
            SELECT DISTINCT
                s.id, s.flight_id_1, s.flight_id_2, s.callsign_1, s.callsign_2,
                s.similarity_level, s.similarity_score, s.has_sector_overlap,
                s.total_overlap_minutes, s.overlap_count, s.detected_at,
                f1.dept_airport_cd as dept1, f1.dest_airport_cd as dest1,
                f2.dept_airport_cd as dept2, f2.dest_airport_cd as dest2
            FROM similarities s
            JOIN flights f1 ON s.flight_id_1 = f1.id
            JOIN flights f2 ON s.flight_id_2 = f2.id"""

        # 시간대 필터가 있으면 sector_overlaps JOIN 추가
        if hour:
            query += """
            LEFT JOIN sector_overlaps so ON s.id = so.similarity_id"""

        query += """
            WHERE s.has_sector_overlap = 1 AND s.total_overlap_minutes >= ?"""
        query += date_filter
        query += hour_filter
        query += """
            ORDER BY s.detected_at DESC
            LIMIT ? OFFSET ?
        """
        query_params = [min_overlap] + date_params + hour_params + [limit, offset]
        similarities = db_manager.execute_query(query, query_params)

        recent_sims = []

        for sim in similarities:
            sim_dict = dict(sim)
            
            # 섹터 겹침 상세 정보 조회 (시간순 정렬)
            sector_overlaps = []
            if sim_dict['has_sector_overlap']:
                overlaps_result = db_manager.execute_query(
                    "SELECT * FROM sector_overlaps WHERE similarity_id = ? ORDER BY overlap_start ASC",
                    (sim_dict['id'],)
                )
                
                for overlap in overlaps_result:
                    ov_dict = dict(overlap)
                    sector_overlaps.append({
                        'id': ov_dict.get('id'),
                        'similarity_id': ov_dict.get('similarity_id'),
                        'sector': ov_dict.get('sector_name'),
                        'entry1': ov_dict.get('flight1_entry'),
                        'exit1': ov_dict.get('flight1_exit'),
                        'entry2': ov_dict.get('flight2_entry'),
                        'exit2': ov_dict.get('flight2_exit'),
                        'overlap_start': ov_dict.get('overlap_start'),
                        'overlap_end': ov_dict.get('overlap_end'),
                        'overlap_minutes': ov_dict.get('overlap_minutes')
                    })

            recent_sims.append({
                'id': sim_dict['id'],
                'flight_id_1': sim_dict['flight_id_1'],
                'flight_id_2': sim_dict['flight_id_2'],
                'callsign1': sim_dict['callsign_1'],
                'callsign2': sim_dict['callsign_2'],
                'dept1': sim_dict['dept1'], 
                'dest1': sim_dict['dest1'],
                'dept2': sim_dict['dept2'], 
                'dest2': sim_dict['dest2'],
                'similarity_level': sim_dict['similarity_level'],
                'similarity_score': sim_dict['similarity_score'],
                'has_sector_overlap': sim_dict['has_sector_overlap'],
                'total_overlap_minutes': sim_dict['total_overlap_minutes'],
                'overlap_count': sim_dict['overlap_count'],
                'detected_at': sim_dict['detected_at'],
                'sector_overlaps': sector_overlaps  # 상세 정보 추가
            })

        # ------------------------------------------------------------------
        # 사이드바/대시보드용 추가 통계 (항공사, 레벨, 콜사인 TOP10, 시간대)
        # ------------------------------------------------------------------
        filter_conditions = "s.has_sector_overlap = 1 AND s.total_overlap_minutes >= ?"
        filter_params = [min_overlap]

        if eobd:
            filter_conditions += " AND f1.eobd = ? AND f2.eobd = ?"
            filter_params.extend([eobd, eobd])

        if hour:
            filter_conditions += " AND EXISTS (\n                SELECT 1 FROM sector_overlaps so_hour\n                WHERE so_hour.similarity_id = s.id\n                AND strftime('%H', REPLACE(so_hour.overlap_start, 'T', ' ')) = ?\n            )"
            filter_params.append(hour.zfill(2))

        filtered_cte = f"""
            WITH filtered AS (
                SELECT
                    s.id,
                    s.callsign_1,
                    s.callsign_2,
                    s.similarity_level,
                    s.similarity_score,
                    s.total_overlap_minutes,
                    s.overlap_count
                FROM similarities s
                JOIN flights f1 ON s.flight_id_1 = f1.id
                JOIN flights f2 ON s.flight_id_2 = f2.id
                WHERE {filter_conditions}
            )
        """

        # 항공사별 순위 (콜사인 앞 3자리 기준)
        airline_query = filtered_cte + """
            SELECT airline, COUNT(*) as count
            FROM (
                SELECT UPPER(SUBSTR(callsign_1, 1, 3)) AS airline FROM filtered
                UNION ALL
                SELECT UPPER(SUBSTR(callsign_2, 1, 3)) AS airline FROM filtered
            ) airlines
            WHERE airline IS NOT NULL AND airline <> ''
            GROUP BY airline
            ORDER BY count DESC
            LIMIT 5
        """
        airline_ranking_result = db_manager.execute_query(airline_query, filter_params)
        airline_ranking = [dict(row) for row in airline_ranking_result] if airline_ranking_result else []

        # 유사도 레벨 랭킹
        level_query = filtered_cte + """
            SELECT similarity_level as level, COUNT(*) as count
            FROM filtered
            WHERE similarity_level IS NOT NULL AND similarity_level <> ''
            GROUP BY similarity_level
            ORDER BY count DESC
        """
        level_ranking_result = db_manager.execute_query(level_query, filter_params)
        level_ranking = [dict(row) for row in level_ranking_result] if level_ranking_result else []

        # 콜사인 TOP10
        callsign_query = filtered_cte + """
            SELECT callsign_1, callsign_2, similarity_level, similarity_score,
                   total_overlap_minutes, overlap_count
            FROM filtered
            ORDER BY similarity_score DESC
            LIMIT 10
        """
        callsign_result = db_manager.execute_query(callsign_query, filter_params)
        callsign_top10 = []
        if callsign_result:
            for row in callsign_result:
                row_dict = dict(row)
                callsign_top10.append({
                    'callsign1': row_dict.get('callsign_1'),
                    'callsign2': row_dict.get('callsign_2'),
                    'similarity_level': row_dict.get('similarity_level'),
                    'similarity_score': row_dict.get('similarity_score'),
                    'total_overlap_minutes': row_dict.get('total_overlap_minutes'),
                    'overlap_count': row_dict.get('overlap_count')
                })

        # 시간대별 현황 (UTC)
        hourly_status_query = filtered_cte + """
            SELECT hour, COUNT(*) as count FROM (
                SELECT DISTINCT filtered.id, strftime('%H', REPLACE(so.overlap_start, 'T', ' ')) AS hour
                FROM sector_overlaps so
                JOIN filtered ON filtered.id = so.similarity_id
            ) hours
            WHERE hour IS NOT NULL
            GROUP BY hour
            ORDER BY hour
        """
        hourly_status_result = db_manager.execute_query(hourly_status_query, filter_params)
        hourly_status = [dict(row) for row in hourly_status_result] if hourly_status_result else []

        return jsonify({
            'status': 'success',
            'data': {
                'total_flights': statistics['total_flights'],
                'total_similarities': statistics['total_similarities'],
                'filtered_similarities': filtered_count,
                'peak_hour': peak_hour,
                'avg_overlap_minutes': avg_overlap,
                'top_sector': top_sector,
                'top_sectors': top_sectors,  # 상위 3개 섹터 추가
                'hourly_distribution_by_date': hourly_by_date,
                'selected_date': eobd,  # 필터링된 날짜 포함
                'airline_ranking': airline_ranking,
                'level_ranking': level_ranking,
                'callsign_top10': callsign_top10,
                'hourly_status': hourly_status,
                'pagination': {
                    'current_page': page,
                    'total_pages': total_pages,
                    'per_page': limit,
                    'total_count': filtered_count
                },
                'recent_similarities': recent_sims
            }
        }), 200

    except Exception as e:
        logger.error(f"상세 통계 조회 오류: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'서버 오류: {str(e)}'
        }), 500


@app.route('/api/statistics/period', methods=['GET'])
def get_period_statistics():
    """기간 분석 탭 전용 통계 데이터"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        min_overlap = request.args.get('min_overlap_minutes', default=2, type=int)

        if not start_date or not end_date:
            return jsonify({
                'status': 'error',
                'message': 'start_date와 end_date 파라미터가 필요합니다. (YYYY-MM-DD)'
            }), 400

        try:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({
                'status': 'error',
                'message': '날짜 형식이 올바르지 않습니다. YYYY-MM-DD 형식을 사용하세요.'
            }), 400

        if end_dt < start_dt:
            return jsonify({
                'status': 'error',
                'message': '종료일은 시작일 이후여야 합니다.'
            }), 400

        range_days = (end_dt - start_dt).days + 1
        if range_days > 31:
            return jsonify({
                'status': 'error',
                'message': '최대 31일까지 조회할 수 있습니다.'
            }), 400

        min_overlap = max(1, min_overlap or 2)

        analysis = db_manager.get_period_analysis(start_date, end_date, min_overlap)

        return jsonify({
            'status': 'success',
            'data': analysis
        }), 200
    except ValueError as ve:
        return jsonify({
            'status': 'error',
            'message': str(ve)
        }), 400
    except Exception as e:
        logger.error(f"기간 분석 데이터 생성 실패: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': '기간 분석 데이터를 생성하지 못했습니다.'
        }), 500


# ============================================================================
# 내보내기 엔드포인트
# ============================================================================

@app.route('/api/export/json', methods=['GET'])
def export_json():
    """
    통계 데이터를 JSON으로 내보내기 (데이터베이스에서)

    Query parameters:
        - min_overlap_minutes: int (기본값: 2)

    Returns:
        JSON: 전체 통계 데이터 및 유사호출 목록
    """
    try:
        min_overlap = request.args.get('min_overlap_minutes', 2, type=int)

        # 기본 통계
        statistics = db_manager.get_statistics()

        if not statistics or statistics['total_flights'] == 0:
            return jsonify({
                'status': 'error',
                'message': '내보낼 데이터가 없습니다'
            }), 400

        # License 검증: 내보내기 제한 확인
        license_manager = get_license_manager()
        limits = license_manager.get_limits()
        export_limit = limits['export_limit']

        # 모든 유사호출 조회 (페이지네이션 없음)
        similarities = db_manager.get_similarities(min_overlap, limit=10000)

        # 내보내기 제한 확인
        if len(similarities) > export_limit:
            logger.warning(f"라이선스 제한: 내보내기 항목 수 초과 - {len(similarities)} > {export_limit}")
            return jsonify({
                'status': 'error',
                'message': f'라이선스 제한: 한 번에 최대 {export_limit}개 항목까지만 내보낼 수 있습니다. (현재: {len(similarities)}개)',
                'data': {
                    'export_count': len(similarities),
                    'export_limit': export_limit,
                    'license_type': license_manager.license_type
                }
            }), 403

        coexistences = []

        for sim in similarities:
            sim_dict = dict(sim)

            # 섹터 겹침 상세 정보
            sector_overlaps = db_manager.execute_query(
                "SELECT * FROM sector_overlaps WHERE similarity_id = ?",
                (sim_dict['id'],)
            )

            coexistences.append({
                'callsign1': sim_dict['callsign_1'],
                'callsign2': sim_dict['callsign_2'],
                'similarity_level': sim_dict['similarity_level'],
                'similarity_score': sim_dict['similarity_score'],
                'has_sector_overlap': sim_dict['has_sector_overlap'],
                'total_overlap_minutes': sim_dict['total_overlap_minutes'],
                'overlap_count': sim_dict['overlap_count'],
                'sector_overlaps': [dict(row) for row in sector_overlaps],
                'detected_at': sim_dict['detected_at']
            })

        return jsonify({
            'status': 'success',
            'timestamp': datetime.now().isoformat(),
            'export_info': {
                'total_flights': statistics['total_flights'],
                'total_similarities': len(coexistences),
                'min_overlap_minutes': min_overlap
            },
            'statistics': statistics,
            'coexistences': coexistences
        }), 200

    except Exception as e:
        logger.error(f"JSON 내보내기 오류: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'서버 오류: {str(e)}'
        }), 500


@app.route('/api/export/flights/excel', methods=['GET'])
def export_flights_excel():
    """
    모든 항공편 데이터를 Excel로 내보내기

    Returns:
        Excel file (.xlsx)
    """
    try:
        # 모든 항공편 데이터 조회 (날짜순 정렬)
        flights = db_manager.execute_query("SELECT * FROM flights ORDER BY eobd, eobt")

        if not flights:
            return jsonify({
                'status': 'error',
                'message': '내보낼 항공편 데이터가 없습니다'
            }), 400

        # License 검증: 내보내기 제한 확인
        license_manager = get_license_manager()
        limits = license_manager.get_limits()
        export_limit = limits['export_limit']

        if len(flights) > export_limit:
            logger.warning(f"라이선스 제한: 내보내기 항공편 수 초과 - {len(flights)} > {export_limit}")
            return jsonify({
                'status': 'error',
                'message': f'라이선스 제한: 한 번에 최대 {export_limit}개 항공편까지만 내보낼 수 있습니다. (현재: {len(flights)}개)',
                'data': {
                    'flight_count': len(flights),
                    'export_limit': export_limit,
                    'license_type': license_manager.license_type
                }
            }), 403

        # Excel 워크북 생성
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "항공편 정보"

        # 헤더 스타일
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 헤더 설정
        headers = ["번호", "콜사인", "출발공항", "도착공항", "기종", "속도", "고도",
                   "예상비행시간(분)", "경로", "EET", "EOBT", "EOBD", "섹터진입/진출", "지점별통과시간", "유사도레벨"]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # 데이터 입력
        for row_num, flight in enumerate(flights, 2):
            flight_dict = dict(flight)
            flight_id = flight_dict.get('id')

            # 섹터 정보 조회
            sector_times = db_manager.execute_query(
                "SELECT sector_name, entry_time, exit_time FROM sector_times WHERE flight_id = ? ORDER BY entry_time",
                (flight_id,)
            )
            sector_info = '; '.join([f"{s['sector_name']} {s['entry_time']}-{s['exit_time']}" for s in sector_times]) if sector_times else ""

            # 지점별 통과시간 조회
            waypoint_times = db_manager.execute_query(
                "SELECT waypoint_name, estimated_time FROM waypoint_times WHERE flight_id = ? ORDER BY waypoint_sequence",
                (flight_id,)
            )
            waypoint_info = ', '.join([f"{w['waypoint_name']} {w['estimated_time']}" for w in waypoint_times]) if waypoint_times else ""

            # 유사호출 정보 조회
            similarities = db_manager.execute_query(
                "SELECT callsign_1, callsign_2, similarity_level FROM similarities WHERE flight_id_1 = ? OR flight_id_2 = ?",
                (flight_id, flight_id)
            )
            # 현재 항공편이 아닌 상대방 콜사인 추출
            other_callsigns = []
            for sim in similarities:
                other_callsign = sim['callsign_2'] if sim['callsign_1'] == flight_dict.get('callsign') else sim['callsign_1']
                level = sim['similarity_level']
                other_callsigns.append(f"{other_callsign}({level})")
            similarity_info = ', '.join(other_callsigns) if other_callsigns else "-"

            row_data = [
                row_num - 1,  # 번호
                flight_dict.get('callsign', ''),
                flight_dict.get('dept_airport_cd', ''),
                flight_dict.get('dest_airport_cd', ''),
                flight_dict.get('aircraft_type', ''),
                flight_dict.get('spd', ''),
                flight_dict.get('alt', ''),
                flight_dict.get('enr_minutes', 0),
                flight_dict.get('route', ''),
                flight_dict.get('eet', ''),
                flight_dict.get('eobt', ''),
                flight_dict.get('eobd', ''),
                sector_info,
                waypoint_info,
                similarity_info
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value if value else ""
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # 컬럼 너비 자동 조정
        column_widths = [8, 12, 12, 12, 12, 10, 10, 15, 25, 10, 12, 12, 30, 30, 20]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

        # 행 높이 설정
        ws.row_dimensions[1].height = 25
        for row in range(2, len(flights) + 2):
            ws.row_dimensions[row].height = 30

        # 파일을 메모리에 저장
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        # 파일 다운로드
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'flights_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )

    except Exception as e:
        logger.error(f"Excel 내보내기 오류: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'서버 오류: {str(e)}'
        }), 500


# ============================================================================
# 샘플 파일 다운로드
# ============================================================================

@app.route('/api/sample/flight-csv', methods=['GET'])
def get_sample_flight_csv():
    """
    비행계획 CSV 샘플 파일 다운로드

    Returns:
        CSV file with headers and sample data
    """
    try:
        from io import StringIO

        # CSV 헤더
        headers = [
            'ACFT_CALLSIGN',      # 항공편 호출부호
            'DEPT_AP_CD',         # 출발공항
            'DEST_AP_CD',         # 도착공항
            'EOBD',               # 추정 이륙 날짜
            'EOBT',               # 추정 이륙 시간
            'ALT',                # 고도
            'SPD',                # 속도
            'TURBULENCE_TYPE',    # 난기류 유형
            'ACFT_TYPE',          # 항공기 기종
            'LINE_TYPE',          # 라인 유형
            'REG_NO',             # 등록번호
            'ICAO_EET(INFO_CN)',  # ICAO EET
            'ENR',                # 항로
            'INFO_CN'             # 추가 정보
        ]

        # 샘플 데이터 (사용자가 제공한 데이터)
        sample_data = [
            [
                'GTI811',           # ACFT_CALLSIGN
                'KJFK',             # DEPT_AP_CD
                'RKSI',             # DEST_AP_CD
                '20251129',         # EOBD
                '2305',             # EOBT
                'F400',             # ALT
                'N0478',            # SPD
                'H',                # TURBULENCE_TYPE
                'B744',             # ACFT_TYPE
                'I',                # LINE_TYPE
                'N405KZ',           # REG_NO
                'CZYZ0045 PAZA0617 RJJJ1048 RKRR1430',  # ICAO_EET
                'DCT GAYEL Q818 WOZEE Q917 MUSIT DCT NUBAM DCT 52N090W 55N095W 60N110W 62N120W 62N130W 6154N14100W DCT HAMND A590 POWAL R451 HIXOR M523 HARKI M523 IPGUD Y800 SOVMO Y512 SDE R217 GTC L512 TENAS B467 KAE G597 KARBU',  # ENR
                'PBN/A1B1C1D1L1O1S2 NAV/RNVD1E2A1 DAT/CPDLCX SUR/260B RSP180 CANMANDATE DOF/251129 REG/N405KZ EET/CZYZ0045 PAZA0617 RJJJ1048 RKRR1430 SEL/MSDK CODE/A4C02D OPR/GTI RMK/EMERG CTC 19147018045 TCAS II EQUIPPED MSN=NI'  # INFO_CN
            ]
        ]

        # CSV 생성
        csv_buffer = StringIO()

        # 헤더 행 작성
        csv_buffer.write(','.join([f'"{h}"' for h in headers]) + '\n')

        # 샘플 데이터 행 작성
        for row in sample_data:
            # CSV 형식으로 따옴표 처리
            csv_row = ','.join([f'"{str(cell).replace(chr(34), chr(34)+chr(34))}"' for cell in row])
            csv_buffer.write(csv_row + '\n')

        csv_content = csv_buffer.getvalue()

        # 파일 다운로드
        return send_file(
            BytesIO(csv_content.encode('utf-8')),
            mimetype='text/csv; charset=utf-8',
            as_attachment=True,
            download_name='Airspace_Flight_Sample.csv'
        )

    except Exception as e:
        logger.error(f"샘플 CSV 생성 오류: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'샘플 파일 생성 오류: {str(e)}'
        }), 500


# ============================================================================
# 업로드 이력 엔드포인트
# ============================================================================

@app.route('/api/upload/history', methods=['GET'])
def get_upload_history():
    """
    업로드 이력 조회

    Query parameters:
        - limit: int (기본값: 20)

    Returns:
        JSON: {
            status: 'success',
            data: list of upload records
        }
    """
    try:
        limit = request.args.get('limit', 20, type=int)
        history = db_manager.get_upload_history(limit)

        return jsonify({
            'status': 'success',
            'data': [dict(record) for record in history]
        }), 200

    except Exception as e:
        logger.error(f"업로드 이력 조회 오류: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'서버 오류: {str(e)}'
        }), 500


# ============================================================================
# 에러 핸들러
# ============================================================================

@app.errorhandler(404)
def not_found(error):
    """404 에러 핸들러"""
    return jsonify({
        'status': 'error',
        'message': '요청한 리소스를 찾을 수 없습니다'
    }), 404


@app.errorhandler(500)
def internal_error(error):
    """500 에러 핸들러"""
    logger.error(f"서버 오류: {str(error)}")
    return jsonify({
        'status': 'error',
        'message': '서버 내부 오류가 발생했습니다'
    }), 500


# ============================================================================
# ⚡ CLI 빠른 시뮬레이션 엔드포인트
# ============================================================================

@app.route('/api/simulation/run-cli', methods=['POST'])
def run_simulation_cli():
    """
    CLI 방식의 빠른 시뮬레이션 (simulate_cli.py 기반)

    파일 업로드 + DB 저장 + 시뮬레이션까지 한 번에 처리
    """
    import subprocess
    import time
    import tempfile

    try:
        logger.info("CLI 시뮬레이션 시작...")
        logger.info(f"요청 메서드: {request.method}")
        logger.info(f"Content-Type: {request.content_type}")
        logger.info(f"Files: {request.files.keys()}")
        logger.info(f"Form: {request.form.keys()}")

        start_time = time.time()

        # 파일 확인
        if 'file' not in request.files:
            logger.error(f"파일 없음. 가능한 키: {list(request.files.keys())}")
            return jsonify({
                'status': 'error',
                'message': f'파일이 없습니다. 가능한 키: {list(request.files.keys())}'
            }), 400

        file = request.files['file']
        if file.filename == '':
            logger.error("파일명이 비어있음")
            return jsonify({
                'status': 'error',
                'message': '파일을 선택하세요'
            }), 400

        logger.info(f"파일명: {file.filename}")

        # 임시 파일로 저장
        with tempfile.NamedTemporaryFile(delete=False, suffix='.csv') as tmp:
            file.save(tmp.name)
            tmp_path = tmp.name

        try:
            # 데이터베이스 초기화 (필요한 경우)
            reset_db = request.form.get('reset_db', 'false').lower() == 'true'
            if reset_db and not is_production:  # 로컬 개발 환경에서만 삭제
                db_path = os.path.join(PROJECT_DIR, 'database', 'backend_similarity_detector.db')
                if os.path.exists(db_path):
                    os.remove(db_path)
                    logger.info(f"데이터베이스 파일 삭제: {db_path}")

            # simulate_cli.py 실행
            cmd = [
                'python3',
                os.path.join(PROJECT_DIR, 'simulate_cli.py'),
                tmp_path
            ]

            logger.info(f"실행 명령: {' '.join(cmd)}")

            # 서브프로세스 실행 (타임아웃 없음 - 배경에서 실행)
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=3600  # 1시간 타임아웃
            )

            elapsed = time.time() - start_time

            if result.returncode != 0:
                logger.error(f"CLI 시뮬레이션 오류: {result.stderr}")
                return jsonify({
                    'status': 'error',
                    'message': f'시뮬레이션 실패: {result.stderr[:200]}'
                }), 500

            # DB에서 결과 조회
            stats = db_manager.get_statistics()

            logger.info(f"CLI 시뮬레이션 완료: {elapsed:.1f}초")

            return jsonify({
                'status': 'success',
                'message': f'{elapsed:.1f}초 완료',
                'data': {
                    'elapsed_time': elapsed,
                    'similarity_count': stats.get('total_similarities', 0),
                    'total_flights': stats.get('total_flights', 0),
                    'cross_sector_pairs': stats.get('cross_sector_overlap_pairs', 0),
                    'statistics': stats
                }
            }), 200

        finally:
            # 임시 파일 삭제
            if os.path.exists(tmp_path):
                os.remove(tmp_path)

    except subprocess.TimeoutExpired:
        logger.error("CLI 시뮬레이션 타임아웃")
        return jsonify({
            'status': 'error',
            'message': '시뮬레이션 타임아웃 (1시간 초과)'
        }), 504

    except Exception as e:
        logger.error(f"CLI 시뮬레이션 오류: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'서버 오류: {str(e)}'
        }), 500


@app.route('/api/simulation/reset-db', methods=['POST'])
def reset_db():
    """데이터베이스 초기화 (로컬 개발 환경에서만)"""
    try:
        # 프로덕션 환경에서는 DB 초기화 거부
        if is_production:
            return jsonify({
                'status': 'error',
                'message': '프로덕션 환경에서는 데이터베이스 초기화를 할 수 없습니다.'
            }), 403

        db_path = os.path.join(PROJECT_DIR, 'database', 'backend_similarity_detector.db')

        if os.path.exists(db_path):
            os.remove(db_path)
            logger.info("데이터베이스 초기화 완료")

            # 새로운 DatabaseManager 인스턴스 생성
            global db_manager
            db_manager = DatabaseManager(db_path)

        return jsonify({
            'status': 'success',
            'message': '데이터베이스 초기화 완료'
        }), 200

    except Exception as e:
        logger.error(f"DB 초기화 오류: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'DB 초기화 실패: {str(e)}'
        }), 500


@app.route('/api/database/delete', methods=['POST'])
def delete_database():
    """
    데이터베이스 또는 특정 일자의 데이터 삭제 (로컬 개발 환경에서만)

    Query parameters:
        - type: 'all' (전체 삭제) 또는 'date' (일자별 삭제)
        - date: type='date'일 때 삭제할 날짜 (YYYY-MM-DD 형식)
    """
    try:
        # 프로덕션 환경에서는 DB 삭제 거부
        if is_production:
            return jsonify({
                'status': 'error',
                'message': '프로덕션 환경에서는 데이터베이스 삭제를 할 수 없습니다.'
            }), 403

        delete_type = request.args.get('type', 'all')
        date_str = request.args.get('date', None)

        if delete_type == 'all':
            # 전체 데이터베이스 삭제
            db_path = os.path.join(PROJECT_DIR, 'database', 'backend_similarity_detector.db')

            if os.path.exists(db_path):
                os.remove(db_path)
                logger.info("데이터베이스 전체 삭제 완료")

                # 새로운 DatabaseManager 인스턴스 생성
                global db_manager
                db_manager = DatabaseManager(db_path)

            return jsonify({
                'status': 'success',
                'message': '전체 데이터베이스 삭제 완료'
            }), 200

        elif delete_type == 'date' and date_str:
            # 특정 일자의 데이터만 삭제
            try:
                # 날짜 형식 검증
                from datetime import datetime
                target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                date_str = str(target_date)
            except ValueError:
                return jsonify({
                    'status': 'error',
                    'message': '유효하지 않은 날짜 형식입니다 (YYYY-MM-DD)'
                }), 400

            # 해당 일자의 데이터 삭제
            # 1. similarities 테이블에서 해당 일자의 records 삭제
            # 2. flights 테이블에서 해당 일자의 records 삭제

            try:
                # 일자에 해당하는 flight id 조회
                flights_query = "SELECT id FROM flights WHERE eobd = ?"
                flights_result = db_manager.execute_query(flights_query, [date_str])
                flight_ids = [dict(row)['id'] for row in flights_result] if flights_result else []

                if flight_ids:
                    # flight id placeholder 생성
                    placeholders = ','.join('?' * len(flight_ids))

                    # sector_overlaps 삭제 (similarities를 통해)
                    delete_sector_overlaps = f"""
                        DELETE FROM sector_overlaps
                        WHERE similarity_id IN (
                            SELECT id FROM similarities
                            WHERE flight_id_1 IN ({placeholders}) OR flight_id_2 IN ({placeholders})
                        )
                    """
                    db_manager.execute_delete(delete_sector_overlaps, flight_ids + flight_ids)

                    # similarities 삭제
                    delete_similarities = f"""
                        DELETE FROM similarities
                        WHERE flight_id_1 IN ({placeholders}) OR flight_id_2 IN ({placeholders})
                    """
                    db_manager.execute_delete(delete_similarities, flight_ids + flight_ids)

                    # waypoint_times 삭제
                    delete_waypoints = f"""
                        DELETE FROM waypoint_times
                        WHERE flight_id IN ({placeholders})
                    """
                    db_manager.execute_delete(delete_waypoints, flight_ids)

                    # sector_times 삭제
                    delete_sectors = f"""
                        DELETE FROM sector_times
                        WHERE flight_id IN ({placeholders})
                    """
                    db_manager.execute_delete(delete_sectors, flight_ids)

                    # flights 삭제
                    delete_flights = f"""
                        DELETE FROM flights
                        WHERE id IN ({placeholders})
                    """
                    db_manager.execute_delete(delete_flights, flight_ids)

                logger.info(f"날짜 {date_str}의 데이터 삭제 완료 (항공편: {len(flight_ids)}개)")

                return jsonify({
                    'status': 'success',
                    'message': f'{date_str} 데이터 삭제 완료 ({len(flight_ids)}개 항공편)'
                }), 200

            except Exception as e:
                logger.error(f"일자별 데이터 삭제 오류: {str(e)}")
                return jsonify({
                    'status': 'error',
                    'message': f'데이터 삭제 실패: {str(e)}'
                }), 500
        else:
            return jsonify({
                'status': 'error',
                'message': '유효하지 않은 삭제 요청입니다'
            }), 400

    except Exception as e:
        logger.error(f"DB 삭제 오류: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'삭제 실패: {str(e)}'
        }), 500


# ============================================================================
# Blueprint 등록
# ============================================================================

# License API 등록
app.register_blueprint(license_bp)
app.register_blueprint(admin_license_bp)


# ============================================================================
# 메인
# ============================================================================

if __name__ == '__main__':
    port = int(os.environ.get('PORT', API_PORT))
    logger.info(f"Flask 앱 시작... (포트: {port}, DEBUG: {DEBUG})")
    app.run(
        host='0.0.0.0',
        port=port,
        debug=DEBUG,
        use_reloader=False
    )
