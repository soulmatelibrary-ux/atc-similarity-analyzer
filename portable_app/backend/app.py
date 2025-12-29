"""
유사호출 감시 시뮬레이션 시스템 - Flask 백엔드 API
"""
import os
import json
import threading
import platform
import signal
from datetime import datetime
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from werkzeug.utils import secure_filename
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from functools import wraps

# 경로 설정
BACKEND_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(BACKEND_DIR)
import sys
sys.path.insert(0, PROJECT_DIR)

from core.similarity_engine import check_similarity, get_risk_level, get_similarity_score
from utils.file_validator import FileValidator
from utils.logger import setup_logger
from utils.similarity_optimizer import SimilarityOptimizer
from utils.sector_parser import parse_sector_times, calculate_sector_overlaps, get_sector_overlap_summary
from utils.constants import API_PORT, FRONTEND_PORT, DEBUG
from database.db_manager import DatabaseManager
from core.flight_service import FlightService

# Flask 앱 초기화
app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['UPLOAD_FOLDER'] = os.path.join(PROJECT_DIR, 'uploads')

# 타임아웃 설정 (유사호출 감지는 최대 45분까지 허용)
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0  # 캐시 비활성화

def timeout_handler(signum, frame):
    raise TimeoutError("Request timeout")

def long_running_task(timeout_seconds=2700):
    """
    긴 작업용 타임아웃 데코레이터 (기본 45분)
    - Unix/Linux: signal.alarm() 사용
    - Windows: threading 기반 타임아웃
    """
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            is_windows = platform.system() == 'Windows'

            if is_windows:
                # Windows: threading 기반 타임아웃
                result_container = [None]
                exception_container = [None]

                def run_function():
                    try:
                        result_container[0] = f(*args, **kwargs)
                    except Exception as e:
                        exception_container[0] = e

                thread = threading.Thread(target=run_function, daemon=True)
                thread.start()
                thread.join(timeout=timeout_seconds)

                if thread.is_alive():
                    logger.warning(f"Request exceeded timeout: {timeout_seconds}s")
                    raise TimeoutError(f"Request timeout after {timeout_seconds}s")

                if exception_container[0]:
                    raise exception_container[0]

                return result_container[0]
            else:
                # Unix/Linux: signal.alarm() 사용
                try:
                    signal.signal(signal.SIGALRM, timeout_handler)
                    signal.alarm(timeout_seconds)
                except Exception as e:
                    logger.warning(f"Timeout handler setup failed: {e}")

                try:
                    result = f(*args, **kwargs)
                finally:
                    try:
                        signal.alarm(0)  # 타이머 취소
                    except Exception as e:
                        logger.warning(f"Timeout cancellation failed: {e}")

                return result

        return wrapper
    return decorator

# CORS 설정 (프론트엔드 포트 기반)
CORS(app, resources={
    r"/api/*": {
        "origins": [
            f"http://localhost:{FRONTEND_PORT}",
            f"http://127.0.0.1:{FRONTEND_PORT}",
            "http://localhost:3000",  # 개발 환경용 추가 포트
        ],
        "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
        "allow_headers": ["Content-Type", "Authorization"],
        "expose_headers": ["Content-Type"],
        "supports_credentials": True,
        "max_age": 3600
    }
})

# 로거 설정
logger = setup_logger(__name__)

# 업로드 폴더 생성
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# 데이터베이스 초기화
db_path = os.path.join(PROJECT_DIR, 'database', 'similarity_detector.db')
db_manager = DatabaseManager(db_path)
flight_service = FlightService(db_manager)

# 상수 정의
ALLOWED_FILE_TYPES = ['csv', 'xlsx', 'xls']
MAX_FILE_SIZE_MB = 16

# 글로벌 상태 (메모리 캐시용)
app_state = {
    'flights': [],
    'coexistences': [],
    'statistics': None
}

# 진행 상황 저장소 (파일 기반 - 안정적)
PROGRESS_DIR = os.path.join(PROJECT_DIR, '.uploads_progress')
os.makedirs(PROGRESS_DIR, exist_ok=True)

def save_upload_progress(process_id, progress_data):
    """진행 상황을 파일에 저장"""
    progress_file = os.path.join(PROGRESS_DIR, f'{process_id}.json')
    try:
        with open(progress_file, 'w', encoding='utf-8') as f:
            json.dump(progress_data, f, ensure_ascii=False)
    except Exception as e:
        logger.error(f"진행 상황 저장 오류: {e}")

def load_upload_progress(process_id):
    """진행 상황을 파일에서 로드"""
    progress_file = os.path.join(PROGRESS_DIR, f'{process_id}.json')
    try:
        if os.path.exists(progress_file):
            with open(progress_file, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception as e:
        logger.error(f"진행 상황 로드 오류: {e}")
    return None

def delete_upload_progress(process_id):
    """진행 상황 파일 삭제"""
    progress_file = os.path.join(PROGRESS_DIR, f'{process_id}.json')
    try:
        if os.path.exists(progress_file):
            os.remove(progress_file)
    except Exception as e:
        logger.error(f"진행 상황 파일 삭제 오류: {e}")


# ============================================================================
# Favicon 엔드포인트 (브라우저 자동 요청 처리)
# ============================================================================

@app.route('/favicon.ico')
def favicon():
    """브라우저의 favicon 요청 처리 (404 방지)"""
    return '', 204  # 204 No Content


@app.route('/')
def index():
    """루트 경로 - 프론트엔드 안내"""
    return """
    <!DOCTYPE html>
    <html>
    <head>
        <title>유사호출 감시 시뮬레이션 시스템 백엔드</title>
        <style>
            body { font-family: sans-serif; text-align: center; padding-top: 50px; }
            .container { max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #ddd; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
            h1 { color: #2c3e50; }
            p { color: #7f8c8d; }
            a { display: inline-block; background-color: #3498db; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px; margin-top: 20px; }
            a:hover { background-color: #2980b9; }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>유사호출 감시 시뮬레이션 시스템 API 서버</h1>
            <p>이곳은 백엔드 API 서버입니다.</p>
            <p>웹 인터페이스를 사용하려면 아래 버튼을 클릭하세요.</p>
            <a href="http://localhost:8000">대시보드로 이동 (프론트엔드)</a>
        </div>
    </body>
    </html>
    """


# ============================================================================
# 헬스체크 엔드포인트
# ============================================================================

@app.route('/api/health', methods=['GET'])
def health_check():
    """
    시스템 헬스 체크

    Returns:
        JSON: {status: 'ok', timestamp: ISO8601}
    """
    return jsonify({
        'status': 'ok',
        'timestamp': datetime.now().isoformat(),
        'version': '1.0.0-alpha',
        'phase': 4
    }), 200


@app.route('/api/similarity-levels', methods=['GET'])
def get_similarity_levels():
    """
    유사도 레벨 및 정의 조회

    Returns:
        JSON: {
            status: 'success',
            data: {
                LEVEL_2-1: {risk: 'HIGH', score: 100, description: '...'},
                ...
            }
        }
    """
    try:
        from utils.constants import SIMILARITY_LEVELS

        return jsonify({
            'status': 'success',
            'data': SIMILARITY_LEVELS
        }), 200
    except Exception as e:
        logger.error(f"유사도 레벨 조회 실패: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'서버 오류: {str(e)}'
        }), 500


# ============================================================================
# 파일 업로드 엔드포인트
# ============================================================================

def _validate_file_extension(filename):
    """
    파일 확장자 검증 (화이트리스트)

    Args:
        filename: 파일명

    Returns:
        tuple: (is_valid: bool, error_message: str)
    """
    if '.' not in filename:
        return False, "파일에 확장자가 없습니다"

    ext = filename.rsplit('.', 1)[1].lower()
    if ext not in ALLOWED_FILE_TYPES:
        return False, f"허용되지 않는 파일 형식입니다. 허용: {', '.join(ALLOWED_FILE_TYPES)}"

    return True, ""


def _validate_file_mime_type(file):
    """
    파일 MIME 타입 검증

    Args:
        file: Werkzeug FileStorage 객체

    Returns:
        tuple: (is_valid: bool, error_message: str)
    """
    allowed_mimes = [
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  # .xlsx
        'application/vnd.ms-excel',  # .xls
        'text/csv',
        'application/csv',
    ]

    mime_type = file.content_type or ''

    # MIME 타입이 명시되지 않은 경우 파일명으로 유추
    if not mime_type or mime_type not in allowed_mimes:
        # 확장자 기반 검증만 수행
        return True, ""

    if mime_type not in allowed_mimes:
        return False, f"허용되지 않는 파일 타입입니다: {mime_type}"

    return True, ""


@app.route('/api/upload/flights', methods=['POST'])
def upload_flights():
    """
    항공편 데이터 파일 업로드

    Request:
        - file: Excel/CSV 파일
        - mode: 'replace' (덮어쓰기) 또는 'append' (누적) - 기본값 'replace'

    Returns:
        JSON: { status, message, data }
    """
    try:
        if 'file' not in request.files:
            return jsonify({
                'status': 'error',
                'message': '파일이 제공되지 않았습니다'
            }), 400

        file = request.files['file']
        mode = request.form.get('mode', 'replace')  # 기본값: 덮어쓰기

        if file.filename == '':
            return jsonify({
                'status': 'error',
                'message': '파일명이 비어있습니다'
            }), 400

        # 1단계: 파일 확장자 검증
        is_valid, error_msg = _validate_file_extension(file.filename)
        if not is_valid:
            logger.warning(f"파일 확장자 검증 실패: {file.filename} - {error_msg}")
            return jsonify({
                'status': 'error',
                'message': error_msg
            }), 400

        # ... (중간 생략 등 기존 코드 유지 가능하므로 아래 로직만 수정) ...
        # 2단계: MIME 타입 검증
        is_valid, error_msg = _validate_file_mime_type(file)
        if not is_valid:
            logger.warning(f"파일 MIME 타입 검증 실패: {file.filename} - {error_msg}")
            return jsonify({
                'status': 'error',
                'message': error_msg
            }), 400

        # 3단계: 파일명 보안 처리
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

        # 4단계: 파일 저장 (검증된 확장자만 저장)
        file.save(filepath)

        # 5단계: 저장된 파일 검증
        validator = FileValidator()
        validation_result = validator.validate_file(filepath)

        if validation_result['status'] == 'invalid':
            os.remove(filepath)  # 저장된 파일 삭제
            logger.warning(f"저장된 파일 검증 실패: {filename}")
            return jsonify({
                'status': 'error',
                'message': '파일 검증 실패',
                'data': {
                    'file_name': filename,
                    'errors': validation_result['errors'],
                    'warnings': validation_result['warnings']
                }
            }), 400

        # 6단계: 데이터 저장 (데이터베이스) - 백그라운드에서 처리
        flights_data = validation_result['data'].to_dict('records')

        # 파일 크기 계산
        file.seek(0, os.SEEK_END)
        file_size = file.tell()
        file.seek(0)

        # 진행 상황 추적용 프로세스 ID 생성
        import uuid
        process_id = str(uuid.uuid4())

        # 백그라운드 스레드에서 처리
        def process_flights_background():
            try:
                # 데이터 처리 모드에 따라 분기
                if mode == 'replace':
                    # 기존 데이터 초기화 (새 파일 업로드 시 깨끗한 상태에서 시작)
                    logger.info("기존 데이터 초기화 중 (덮어쓰기 모드)...")
                    db_manager.execute_query("DELETE FROM sector_overlaps")
                    db_manager.execute_query("DELETE FROM similarities")
                    db_manager.execute_query("DELETE FROM sector_times")
                    db_manager.execute_query("DELETE FROM flights")
                    logger.info("기존 데이터 초기화 완료")
                    
                    # 메모리 캐시 초기화
                    app_state['flights'] = []
                else:
                    logger.info("데이터 누적 모드로 업로드 진행...")

                service_result = flight_service.process_and_save_flights(
                    flights_data,
                    filename,
                    file_size,
                    process_id=process_id
                )
                
                # 메모리 캐시 업데이트 (누적)
                if app_state['flights'] is None:
                    app_state['flights'] = []
                app_state['flights'].extend(flights_data)
                
                logger.info(f"백그라운드 파일 처리 완료: {filename} ({len(flights_data)}건) - 모드: {mode}")
            except Exception as e:
                logger.error(f"백그라운드 파일 처리 오류: {str(e)}")
                from core.flight_service import upload_progress
                upload_progress[process_id] = {
                    'total': len(flights_data),
                    'processed': 0,
                    'stage': f'오류: {str(e)}',
                    'percent': 0
                }

        # 백그라운드 스레드 시작 (데몬 스레드)
        bg_thread = threading.Thread(target=process_flights_background, daemon=True)
        bg_thread.start()

        logger.info(f"파일 업로드 요청 접수: {filename} ({len(flights_data)}건) - process_id: {process_id}")

        # 즉시 응답 (백그라운드에서 처리 중)
        return jsonify({
            'status': 'success',
            'message': '파일 업로드 시작 - 백그라운드에서 처리 중입니다',
            'data': {
                'file_name': filename,
                'record_count': len(flights_data),
                'process_id': process_id,
                'errors': validation_result['errors'],
                'warnings': validation_result['warnings']
            }
        }), 200

    except Exception as e:
        logger.error(f"파일 업로드 오류: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'서버 오류: {str(e)}'
        }), 500


# ============================================================================
# 업로드 진행 상황 조회 엔드포인트
# ============================================================================

@app.route('/api/upload/progress/<process_id>', methods=['GET'])
def get_upload_progress(process_id):
    """
    파일 업로드 진행 상황 조회

    Args:
        process_id: 진행 상황 추적 ID

    Returns:
        JSON: {
            'status': 'in_progress'/'completed'/'error',
            'percent': int (0-100),
            'stage': str,
            'total': int,
            'processed': int
        }
    """
    try:
        # 파일에서 진행 상황 로드
        progress_data = load_upload_progress(process_id)

        if progress_data is None:
            return jsonify({
                'status': 'not_found',
                'percent': 0,
                'stage': 'Unknown process'
            }), 404

        return jsonify({
            'status': 'in_progress' if progress_data['percent'] < 100 else 'completed',
            'percent': progress_data['percent'],
            'stage': progress_data['stage'],
            'total': progress_data['total'],
            'processed': progress_data['processed']
        }), 200

    except Exception as e:
        logger.error(f"진행 상황 조회 오류: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


# ============================================================================
# 유사호출 판정 엔드포인트
# ============================================================================

@app.route('/api/similarity/check', methods=['POST'])
def check_similarity_api():
    """
    두 콜사인의 유사도 판정

    Request:
        JSON: {
            callsign1: str,
            callsign2: str
        }

    Returns:
        JSON: {
            callsign1: str,
            callsign2: str,
            similarity_level: str,
            risk_level: str,
            score: int,
            edit_distance: int
        }
    """
    try:
        data = request.get_json()

        if not data or 'callsign1' not in data or 'callsign2' not in data:
            return jsonify({
                'status': 'error',
                'message': '필수 파라미터: callsign1, callsign2'
            }), 400

        callsign1 = data['callsign1']
        callsign2 = data['callsign2']

        # 유사도 판정
        level, ed = check_similarity(callsign1, callsign2)
        risk = get_risk_level(level)
        score = get_similarity_score(level)

        logger.info(f"유사호출 판정: {callsign1} vs {callsign2} → {level}")

        return jsonify({
            'status': 'success',
            'data': {
                'callsign1': callsign1,
                'callsign2': callsign2,
                'similarity_level': level,
                'risk_level': risk,
                'score': score,
                'edit_distance': ed
            }
        }), 200

    except Exception as e:
        logger.error(f"유사호출 판정 오류: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'서버 오류: {str(e)}'
        }), 500


# ============================================================================
# 시뮬레이션 실행 엔드포인트
# ============================================================================

@app.route('/api/simulation/run', methods=['POST'])
def run_simulation():
    """
    유사호출 감지 및 공존 계산 시뮬레이션 실행

    Request:
        JSON: {
            filters: {
                min_coexist_minutes: float (선택),
                max_coexist_minutes: float (선택),
                sectors: list (선택),
                risk_levels: list (선택)
            }
        }

    Returns:
        JSON: {
            status: 'success'|'error',
            data: {
                total_events: int,
                coexistences: list,
                statistics: dict,
                estimated_time_seconds: int,
                estimated_time_minutes: float,
                total_pairs: int
            }
        }
    """
    try:
        data = request.get_json() or {}

        # 최소 공존시간 필터 (기본값: 2분)
        min_overlap_minutes = data.get('min_overlap_minutes', 2)

        # 항공편 수 확인 및 예상 시간 계산
        flights_count_result = db_manager.execute_query("SELECT COUNT(*) as count FROM flights")
        flights_count = dict(flights_count_result[0])['count'] if flights_count_result else 0

        # 예상 시간 계산 (항공편 1000개 기준: 약 12분)
        # 총 쌍의 개수: N * (N-1) / 2
        total_pairs = max(0, flights_count * (flights_count - 1) // 2) if flights_count > 0 else 0

        # 경험적 데이터: 1000개 항공편 ≈ ~1.1M 쌍 ≈ 12분 (약 0.0066ms/쌍)
        # 1481개 항공편 ≈ ~1.1M 쌍 ≈ 18분
        avg_time_per_pair_ms = 0.0065  # milliseconds
        estimated_time_seconds = max(30, int((total_pairs * avg_time_per_pair_ms) / 1000))
        estimated_time_minutes = estimated_time_seconds / 60

        logger.info(f"⏱️  시뮬레이션 예상 시간: {flights_count}개 항공편 → {total_pairs:,}개 쌍 → 약 {estimated_time_minutes:.1f}분 ({estimated_time_seconds}초)")

        # 기존 유사호출 결과 초기화 (새로운 계산을 위해 깨끗한 상태)
        logger.info("기존 유사호출 결과 초기화 중...")
        db_manager.execute_query("DELETE FROM sector_overlaps")
        db_manager.execute_query("DELETE FROM similarities")
        logger.info("기존 유사호출 결과 초기화 완료")

        # 데이터베이스에서 유사호출 감지 실행
        service_result = flight_service.detect_similarities(min_overlap_minutes)

        if service_result['status'] == 'error':
            return jsonify({
                'status': 'error',
                'message': service_result['message']
            }), 400

        # 데이터베이스에서 유사호출 결과 조회
        similarities = db_manager.get_similarities(min_overlap_minutes, limit=100)

        # 결과를 API 응답 형식으로 변환
        coexistences = []
        for sim in similarities:
            similarity_dict = dict(sim)

            # 섹터 겹침 상세 정보 조회
            sector_overlaps_result = db_manager.execute_query(
                "SELECT * FROM sector_overlaps WHERE similarity_id = ?",
                (similarity_dict['id'],)
            )
            sector_overlaps_raw = [dict(row) for row in sector_overlaps_result]

            # 프론트엔드 기대 형식으로 변환
            sector_overlaps = []
            for overlap in sector_overlaps_raw:
                sector_overlaps.append({
                    'id': overlap.get('id'),
                    'similarity_id': overlap.get('similarity_id'),
                    'sector': overlap.get('sector_name'),  # sector_name -> sector
                    'entry1': overlap.get('flight1_entry'),
                    'exit1': overlap.get('flight1_exit'),
                    'entry2': overlap.get('flight2_entry'),
                    'exit2': overlap.get('flight2_exit'),
                    'overlap_start': overlap.get('overlap_start'),
                    'overlap_end': overlap.get('overlap_end'),
                    'overlap_minutes': overlap.get('overlap_minutes')
                })

            coexistence_entry = {
                'id': similarity_dict['id'],
                'flight_id_1': similarity_dict['flight_id_1'],  # ✅ 프론트엔드 필터링용
                'flight_id_2': similarity_dict['flight_id_2'],  # ✅ 프론트엔드 필터링용
                'callsign1': similarity_dict['callsign_1'],
                'callsign2': similarity_dict['callsign_2'],
                'similarity_level': similarity_dict['similarity_level'],  # ✅ 유사도 레벨
                'similarity_score': similarity_dict.get('similarity_score', 0),
                'sector_overlaps': sector_overlaps,
                'sector_summary': {
                    'total_overlaps': similarity_dict['overlap_count'],
                    'total_overlap_minutes': similarity_dict['total_overlap_minutes']
                }
            }

            coexistences.append(coexistence_entry)

        # 캐시된 통계 조회 또는 새로 계산
        cached_stats = db_manager.get_cached_statistics()
        if cached_stats:
            summary = cached_stats
        else:
            summary = db_manager.get_statistics()

        # 메모리 캐시도 업데이트 (백엔드 호환성)
        app_state.update({
            'coexistences': coexistences,
            'statistics': summary
        })

        logger.info(f"시뮬레이션 완료: {len(coexistences)}개 유사호출 감지")

        return jsonify({
            'status': 'success',
            'data': {
                'total_events': len(coexistences),
                'coexistences': coexistences[:100],  # 처음 100개만
                'statistics': summary,
                'estimated_time_seconds': estimated_time_seconds,
                'estimated_time_minutes': round(estimated_time_minutes, 1),
                'total_pairs': total_pairs
            }
        }), 200

    except Exception as e:
        logger.error(f"시뮬레이션 오류: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'서버 오류: {str(e)}'
        }), 500


# ============================================================================
# 통계 조회 엔드포인트
# ============================================================================

@app.route('/api/statistics/summary', methods=['GET'])
def get_statistics_summary():
    """
    통계 요약 조회 (데이터베이스에서)

    Returns:
        JSON: {
            status: 'success'|'error',
            data: {
                total_flights: int,
                total_similarities: int,
                similarities_with_overlap: int,
                cross_sector_overlap_pairs: int,
                level_distribution: dict,
                sector_statistics: list,
                overlap_time_distribution: dict,
                overlap_count_distribution: dict
            }
        }
    """
    try:
        # 캐시된 통계 조회 또는 새로 계산
        cached_stats = db_manager.get_cached_statistics()
        if cached_stats:
            statistics = cached_stats
        else:
            statistics = db_manager.get_statistics()

        if not statistics:
            return jsonify({
                'status': 'error',
                'message': '통계를 계산할 데이터가 없습니다'
            }), 400

        return jsonify({
            'status': 'success',
            'data': statistics,
            'cached': cached_stats is not None
        }), 200

    except Exception as e:
        logger.error(f"통계 조회 오류: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'서버 오류: {str(e)}'
        }), 500


@app.route('/api/flights/dates', methods=['GET'])
def get_available_dates():
    """
    사용 가능한 항공편 날짜 목록 조회

    Returns:
        JSON: {
            status: 'success',
            data: list of dates (YYYY-MM-DD format, sorted descending)
        }
    """
    try:
        result = db_manager.execute_query("SELECT DISTINCT eobd FROM flights ORDER BY eobd DESC")
        dates = [row[0] for row in result] if result else []

        return jsonify({
            'status': 'success',
            'data': dates
        })
    except Exception as e:
        logger.error(f"사용 가능한 날짜 조회 실패: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'서버 오류: {str(e)}'
        }), 500


@app.route('/api/flights/all', methods=['GET'])
def get_all_flights():
    """
    모든 항공편 조회 (섹터 정보 포함)

    Query parameters:
        - page: int (기본값: 1)
        - limit: int (기본값: 50)
        - eobd: str (선택사항: YYYY-MM-DD 형식의 날짜 필터)

    Returns:
        JSON: {
            status: 'success',
            data: list of flights with sector info,
            pagination: {total, page, limit, total_pages},
            selected_date: str (필터된 날짜)
        }
    """
    try:
        page = request.args.get('page', 1, type=int)
        limit = request.args.get('limit', 50, type=int)
        eobd = request.args.get('eobd', None)  # 날짜 필터
        offset = (page - 1) * limit

        # 날짜 필터 조건 구성
        date_filter = ""
        date_params = []
        if eobd:
            date_filter = " WHERE eobd = ?"
            date_params = [eobd]

        # 전체 항공편 수 (날짜 필터 적용)
        count_query = f"SELECT COUNT(*) FROM flights{date_filter}"
        total_result = db_manager.execute_query(count_query, date_params)
        total = total_result[0][0] if total_result else 0
        total_pages = (total + limit - 1) // limit

        # 페이지네이션 적용된 항공편 조회 (날짜 필터 적용, 날짜순 정렬)
        query = f"SELECT * FROM flights{date_filter} ORDER BY eobd, eobt LIMIT ? OFFSET ?"
        flights = db_manager.execute_query(
            query,
            date_params + [limit, offset]
        )

        flights_list = []
        for f in flights:
            f_dict = dict(f)

            # 각 항공편의 섹터 통과 시간 조회
            sector_times = db_manager.execute_query(
                "SELECT sector_name, entry_time, exit_time FROM sector_times WHERE flight_id = ? ORDER BY entry_time",
                (f_dict['id'],)
            )
            f_dict['sectors'] = [dict(s) for s in sector_times]

            # 각 항공편의 지점별 통과시간 조회
            waypoint_times = db_manager.execute_query(
                "SELECT waypoint_name, waypoint_sequence, estimated_time, actual_time FROM waypoint_times WHERE flight_id = ? ORDER BY waypoint_sequence",
                (f_dict['id'],)
            )
            f_dict['waypoints'] = [dict(w) for w in waypoint_times]

            # 각 항공편의 유사호출 정보 조회
            similarities = db_manager.execute_query(
                "SELECT id, flight_id_1, flight_id_2, callsign_1, callsign_2, similarity_level, similarity_score, has_sector_overlap, total_overlap_minutes FROM similarities WHERE flight_id_1 = ? OR flight_id_2 = ? ORDER BY detected_at DESC",
                (f_dict['id'], f_dict['id'])
            )

            similarity_info = []
            for sim in similarities:
                sim_dict = dict(sim)
                # 상대방 항공기 정보 추출
                other_flight_id = sim_dict['flight_id_2'] if sim_dict['flight_id_1'] == f_dict['id'] else sim_dict['flight_id_1']
                other_callsign = sim_dict['callsign_2'] if sim_dict['flight_id_1'] == f_dict['id'] else sim_dict['callsign_1']

                similarity_info.append({
                    'other_flight_id': other_flight_id,
                    'other_callsign': other_callsign,
                    'similarity_level': sim_dict['similarity_level'],
                    'similarity_score': sim_dict['similarity_score'],
                    'has_sector_overlap': sim_dict['has_sector_overlap'],  # 실제 공존 여부
                    'total_overlap_minutes': sim_dict['total_overlap_minutes']  # 공존 시간
                })

            f_dict['similarities'] = similarity_info
            flights_list.append(f_dict)

        return jsonify({
            'status': 'success',
            'data': flights_list,
            'pagination': {
                'total': total,
                'page': page,
                'limit': limit,
                'total_pages': total_pages
            }
        }), 200

    except Exception as e:
        logger.error(f"전체 항공편 조회 오류: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'서버 오류: {str(e)}'
        }), 500


@app.route('/api/flights/pair-details', methods=['GET'])
def get_flight_pair_details():
    """
    두 항공편의 상세 정보 조회 (팝업용)
    """
    try:
        flight_id_1 = request.args.get('flight_id_1')
        flight_id_2 = request.args.get('flight_id_2')

        if not flight_id_1 or not flight_id_2:
            return jsonify({'status': 'error', 'message': '항공편 ID가 필요합니다'}), 400

        result = {}
        
        for i, fid in enumerate([flight_id_1, flight_id_2], 1):
            # 기본 항공편 정보
            flight = db_manager.execute_query(
                "SELECT * FROM flights WHERE id = ?", (fid,)
            )
            
            if not flight:
                continue

            f_dict = dict(flight[0])
            
            # 섹터 통과 시간 조회
            sector_times = db_manager.execute_query(
                "SELECT sector_name, entry_time, exit_time FROM sector_times WHERE flight_id = ? ORDER BY entry_time",
                (fid,)
            )

            # 지점별 통과시간 조회
            waypoint_times = db_manager.execute_query(
                "SELECT waypoint_name, waypoint_sequence, estimated_time, actual_time FROM waypoint_times WHERE flight_id = ? ORDER BY waypoint_sequence",
                (fid,)
            )

            result[f'flight{i}'] = {
                'info': f_dict,
                'sectors': [dict(s) for s in sector_times],
                'waypoints': [dict(w) for w in waypoint_times]
            }

        return jsonify({
            'status': 'success',
            'data': result
        })

    except Exception as e:
        logger.error(f"상세 조회 오류: {str(e)}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/statistics/detailed', methods=['GET'])
def get_statistics_detailed():
    """
    상세 통계 조회 (데이터베이스에서)
    상세 통계 조회 (데이터베이스에서) - 페이징 및 필터 적용

    Query parameters:
        - min_overlap_minutes: int (기본값: 2)
        - page: int (기본값: 1)
        - limit: int (기본값: 20)
        - eobd: str (선택, YYYY-MM-DD 형식) - 특정 날짜로 필터링

    Returns:
        JSON: {
            status: 'success',
            data: {
                total_flights: int,
                total_similarities: int,          # 전체 유사호출 수 (필터 무관)
                filtered_similarities: int,       # 필터 적용된 유사호출 수
                pagination: {
                    current_page: int,
                    total_pages: int,
                    per_page: int,
                    total_count: int
                },
                recent_similarities: list,
                selected_date: str or null        # 필터링된 날짜 (있으면)
            }
        }
    """
    try:
        min_overlap = request.args.get('min_overlap_minutes', 2, type=int)
        page = request.args.get('page', 1, type=int)
        limit = request.args.get('limit', 20, type=int)
        eobd = request.args.get('eobd', None)  # 날짜 필터 (YYYY-MM-DD)
        hour = request.args.get('hour', None)  # 시간대 필터 (HH 형식, 00-23)

        offset = (page - 1) * limit

        # 날짜 필터 조건 구성
        date_filter = ""
        date_params = []
        if eobd:
            date_filter = " AND f1.eobd = ? AND f2.eobd = ?"
            date_params = [eobd, eobd]

        # 시간대 필터 조건 구성
        hour_filter = ""
        hour_params = []
        if hour:
            hour_filter = " AND strftime('%H', REPLACE(so.overlap_start, 'T', ' ')) = ?"
            hour_params = [hour]

        # 기본 통계 (날짜 필터 적용)
        if eobd:
            # 날짜별 통계 계산
            total_flights_query = "SELECT COUNT(DISTINCT id) FROM flights WHERE eobd = ?"
            total_flights_result = db_manager.execute_query(total_flights_query, [eobd])
            total_flights = total_flights_result[0][0] if total_flights_result else 0

            total_similarities_query = "SELECT COUNT(*) FROM similarities s JOIN flights f1 ON s.flight_id_1 = f1.id JOIN flights f2 ON s.flight_id_2 = f2.id WHERE f1.eobd = ? AND f2.eobd = ?"
            total_similarities_result = db_manager.execute_query(total_similarities_query, [eobd, eobd])
            total_similarities = total_similarities_result[0][0] if total_similarities_result else 0

            statistics = {
                'total_flights': total_flights,
                'total_similarities': total_similarities,
                'sector_statistics': []
            }
        else:
            # 전체 통계
            statistics = db_manager.get_statistics()

        # 추가 통계: 가장 붐비는 시간대 (Peak Time) - 필터 적용
        peak_time_query = """
            SELECT strftime('%H', REPLACE(so.overlap_start, 'T', ' ')) as hour, COUNT(*) as count
            FROM sector_overlaps so
            JOIN similarities s ON so.similarity_id = s.id
            JOIN flights f1 ON s.flight_id_1 = f1.id
            WHERE s.has_sector_overlap = 1 AND so.overlap_minutes >= ?"""
        peak_params = [min_overlap]

        # 날짜 필터 적용
        if eobd:
            peak_time_query += " AND f1.eobd = ?"
            peak_params.append(eobd)

        peak_time_query += " GROUP BY hour ORDER BY count DESC LIMIT 1"

        peak_result = db_manager.execute_query(peak_time_query, peak_params)
        peak_hour = peak_result[0]['hour'] if peak_result else "-"

        # 추가 통계: 평균 겹침 시간 (필터 적용)
        avg_query = """SELECT AVG(so.overlap_minutes) as avg_min
                      FROM sector_overlaps so
                      JOIN similarities s ON so.similarity_id = s.id
                      JOIN flights f1 ON s.flight_id_1 = f1.id
                      WHERE s.has_sector_overlap = 1 AND so.overlap_minutes >= ?"""
        avg_params = [min_overlap]
        if eobd:
            avg_query += " AND f1.eobd = ?"
            avg_params.append(eobd)

        avg_result = db_manager.execute_query(avg_query, avg_params)
        avg_overlap = round(dict(avg_result[0])['avg_min'], 1) if avg_result and dict(avg_result[0])['avg_min'] else 0

        # 추가 통계: 최다 빈도 섹터 (상위 3개) - 필터 적용
        top_sectors_query = """
            SELECT so.sector_name, COUNT(*) as count
            FROM sector_overlaps so
            JOIN similarities s ON so.similarity_id = s.id
            JOIN flights f1 ON s.flight_id_1 = f1.id
            WHERE s.has_sector_overlap = 1 AND so.overlap_minutes >= ?"""
        top_sectors_params = [min_overlap]
        if eobd:
            top_sectors_query += " AND f1.eobd = ?"
            top_sectors_params.append(eobd)
        top_sectors_query += " GROUP BY so.sector_name ORDER BY count DESC LIMIT 3"

        top_sectors_result = db_manager.execute_query(top_sectors_query, top_sectors_params)
        top_sectors = [dict(row) for row in top_sectors_result] if top_sectors_result else []
        top_sector = top_sectors[0]['sector_name'] if top_sectors else "-"

        # 추가 통계: 일자별/시간대별 분포 (차트용)
        # eobd를 기준으로 시간대별 분포 계산 - 필터 적용
        hourly_query = """SELECT f1.eobd as date, strftime('%H', REPLACE(so.overlap_start, 'T', ' ')) as hour, COUNT(*) as count
            FROM sector_overlaps so
            JOIN similarities s ON so.similarity_id = s.id
            JOIN flights f1 ON s.flight_id_1 = f1.id
            WHERE s.has_sector_overlap = 1 AND so.overlap_minutes >= ?"""
        hourly_params = [min_overlap]

        # 날짜 필터 적용
        if eobd:
            hourly_query += " AND f1.eobd = ?"
            hourly_params.append(eobd)

        # 시간대 필터 적용
        if hour:
            hourly_query += " AND strftime('%H', REPLACE(so.overlap_start, 'T', ' ')) = ?"
            hourly_params.append(hour.zfill(2))

        hourly_query += " GROUP BY f1.eobd, hour ORDER BY f1.eobd, hour"

        hourly_results = db_manager.execute_query(hourly_query, hourly_params)
        
        # 데이터 구조 변환: { '2024-05-01': { '00': 1, '01': 5 ... }, ... }
        hourly_by_date = {}
        for row in hourly_results:
            d = dict(row)
            date_str = d['date']
            hour_str = d['hour']
            count = d['count']
            
            if date_str not in hourly_by_date:
                hourly_by_date[date_str] = {}
            hourly_by_date[date_str][hour_str] = count

        # 필터링된 총 개수 조회
        count_query = """
            SELECT COUNT(DISTINCT s.id)
            FROM similarities s
            JOIN flights f1 ON s.flight_id_1 = f1.id
            JOIN flights f2 ON s.flight_id_2 = f2.id"""

        # 시간대 필터가 있으면 sector_overlaps JOIN 추가
        if hour:
            count_query += """
            LEFT JOIN sector_overlaps so ON s.id = so.similarity_id"""

        count_query += """
            WHERE s.has_sector_overlap = 1 AND s.total_overlap_minutes >= ?"""
        count_query += date_filter
        count_query += hour_filter

        count_params = [min_overlap] + date_params + hour_params
        result = db_manager.execute_query(count_query, count_params)
        filtered_count = result[0][0] if result else 0

        total_pages = (filtered_count + limit - 1) // limit

        # 페이징 적용된 목록 조회
        query = """
            SELECT DISTINCT
                s.id, s.flight_id_1, s.flight_id_2, s.callsign_1, s.callsign_2,
                s.similarity_level, s.similarity_score, s.has_sector_overlap,
                s.total_overlap_minutes, s.overlap_count, s.detected_at,
                f1.dept_airport_cd as dept1, f1.dest_airport_cd as dest1,
                f2.dept_airport_cd as dept2, f2.dest_airport_cd as dest2
            FROM similarities s
            JOIN flights f1 ON s.flight_id_1 = f1.id
            JOIN flights f2 ON s.flight_id_2 = f2.id"""

        # 시간대 필터가 있으면 sector_overlaps JOIN 추가
        if hour:
            query += """
            LEFT JOIN sector_overlaps so ON s.id = so.similarity_id"""

        query += """
            WHERE s.has_sector_overlap = 1 AND s.total_overlap_minutes >= ?"""
        query += date_filter
        query += hour_filter
        query += """
            ORDER BY s.detected_at DESC
            LIMIT ? OFFSET ?
        """
        query_params = [min_overlap] + date_params + hour_params + [limit, offset]
        similarities = db_manager.execute_query(query, query_params)

        recent_sims = []

        for sim in similarities:
            sim_dict = dict(sim)
            
            # 섹터 겹침 상세 정보 조회 (시간순 정렬)
            sector_overlaps = []
            if sim_dict['has_sector_overlap']:
                overlaps_result = db_manager.execute_query(
                    "SELECT * FROM sector_overlaps WHERE similarity_id = ? ORDER BY overlap_start ASC",
                    (sim_dict['id'],)
                )
                
                for overlap in overlaps_result:
                    ov_dict = dict(overlap)
                    sector_overlaps.append({
                        'id': ov_dict.get('id'),
                        'similarity_id': ov_dict.get('similarity_id'),
                        'sector': ov_dict.get('sector_name'),
                        'entry1': ov_dict.get('flight1_entry'),
                        'exit1': ov_dict.get('flight1_exit'),
                        'entry2': ov_dict.get('flight2_entry'),
                        'exit2': ov_dict.get('flight2_exit'),
                        'overlap_start': ov_dict.get('overlap_start'),
                        'overlap_end': ov_dict.get('overlap_end'),
                        'overlap_minutes': ov_dict.get('overlap_minutes')
                    })

            recent_sims.append({
                'id': sim_dict['id'],
                'flight_id_1': sim_dict['flight_id_1'],
                'flight_id_2': sim_dict['flight_id_2'],
                'callsign1': sim_dict['callsign_1'],
                'callsign2': sim_dict['callsign_2'],
                'dept1': sim_dict['dept1'], 
                'dest1': sim_dict['dest1'],
                'dept2': sim_dict['dept2'], 
                'dest2': sim_dict['dest2'],
                'similarity_level': sim_dict['similarity_level'],
                'similarity_score': sim_dict['similarity_score'],
                'has_sector_overlap': sim_dict['has_sector_overlap'],
                'total_overlap_minutes': sim_dict['total_overlap_minutes'],
                'overlap_count': sim_dict['overlap_count'],
                'detected_at': sim_dict['detected_at'],
                'sector_overlaps': sector_overlaps  # 상세 정보 추가
            })

        return jsonify({
            'status': 'success',
            'data': {
                'total_flights': statistics['total_flights'],
                'total_similarities': statistics['total_similarities'],
                'filtered_similarities': filtered_count,
                'peak_hour': peak_hour,
                'avg_overlap_minutes': avg_overlap,
                'top_sector': top_sector,
                'top_sectors': top_sectors,  # 상위 3개 섹터 추가
                'hourly_distribution_by_date': hourly_by_date,
                'selected_date': eobd,  # 필터링된 날짜 포함
                'pagination': {
                    'current_page': page,
                    'total_pages': total_pages,
                    'per_page': limit,
                    'total_count': filtered_count
                },
                'recent_similarities': recent_sims
            }
        }), 200

    except Exception as e:
        logger.error(f"상세 통계 조회 오류: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'서버 오류: {str(e)}'
        }), 500


# ============================================================================
# 내보내기 엔드포인트
# ============================================================================

@app.route('/api/export/json', methods=['GET'])
def export_json():
    """
    통계 데이터를 JSON으로 내보내기 (데이터베이스에서)

    Query parameters:
        - min_overlap_minutes: int (기본값: 2)

    Returns:
        JSON: 전체 통계 데이터 및 유사호출 목록
    """
    try:
        min_overlap = request.args.get('min_overlap_minutes', 2, type=int)

        # 기본 통계
        statistics = db_manager.get_statistics()

        if not statistics or statistics['total_flights'] == 0:
            return jsonify({
                'status': 'error',
                'message': '내보낼 데이터가 없습니다'
            }), 400

        # 모든 유사호출 조회 (페이지네이션 없음)
        similarities = db_manager.get_similarities(min_overlap, limit=10000)
        coexistences = []

        for sim in similarities:
            sim_dict = dict(sim)

            # 섹터 겹침 상세 정보
            sector_overlaps = db_manager.execute_query(
                "SELECT * FROM sector_overlaps WHERE similarity_id = ?",
                (sim_dict['id'],)
            )

            coexistences.append({
                'callsign1': sim_dict['callsign_1'],
                'callsign2': sim_dict['callsign_2'],
                'similarity_level': sim_dict['similarity_level'],
                'similarity_score': sim_dict['similarity_score'],
                'has_sector_overlap': sim_dict['has_sector_overlap'],
                'total_overlap_minutes': sim_dict['total_overlap_minutes'],
                'overlap_count': sim_dict['overlap_count'],
                'sector_overlaps': [dict(row) for row in sector_overlaps],
                'detected_at': sim_dict['detected_at']
            })

        return jsonify({
            'status': 'success',
            'timestamp': datetime.now().isoformat(),
            'export_info': {
                'total_flights': statistics['total_flights'],
                'total_similarities': len(coexistences),
                'min_overlap_minutes': min_overlap
            },
            'statistics': statistics,
            'coexistences': coexistences
        }), 200

    except Exception as e:
        logger.error(f"JSON 내보내기 오류: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'서버 오류: {str(e)}'
        }), 500


@app.route('/api/export/flights/excel', methods=['GET'])
def export_flights_excel():
    """
    모든 항공편 데이터를 Excel로 내보내기

    Returns:
        Excel file (.xlsx)
    """
    try:
        # 모든 항공편 데이터 조회 (날짜순 정렬)
        flights = db_manager.execute_query("SELECT * FROM flights ORDER BY eobd, eobt")

        if not flights:
            return jsonify({
                'status': 'error',
                'message': '내보낼 항공편 데이터가 없습니다'
            }), 400

        # Excel 워크북 생성
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "항공편 정보"

        # 헤더 스타일
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 헤더 설정
        headers = ["번호", "콜사인", "출발공항", "도착공항", "기종", "속도", "고도",
                   "예상비행시간(분)", "경로", "EET", "EOBT", "EOBD", "섹터진입/진출", "지점별통과시간", "유사도레벨"]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # 데이터 입력
        for row_num, flight in enumerate(flights, 2):
            flight_dict = dict(flight)
            flight_id = flight_dict.get('id')

            # 섹터 정보 조회
            sector_times = db_manager.execute_query(
                "SELECT sector_name, entry_time, exit_time FROM sector_times WHERE flight_id = ? ORDER BY entry_time",
                (flight_id,)
            )
            sector_info = '; '.join([f"{s['sector_name']} {s['entry_time']}-{s['exit_time']}" for s in sector_times]) if sector_times else ""

            # 지점별 통과시간 조회
            waypoint_times = db_manager.execute_query(
                "SELECT waypoint_name, estimated_time FROM waypoint_times WHERE flight_id = ? ORDER BY waypoint_sequence",
                (flight_id,)
            )
            waypoint_info = ', '.join([f"{w['waypoint_name']} {w['estimated_time']}" for w in waypoint_times]) if waypoint_times else ""

            # 유사호출 정보 조회
            similarities = db_manager.execute_query(
                "SELECT callsign_1, callsign_2, similarity_level FROM similarities WHERE flight_id_1 = ? OR flight_id_2 = ?",
                (flight_id, flight_id)
            )
            # 현재 항공편이 아닌 상대방 콜사인 추출
            other_callsigns = []
            for sim in similarities:
                other_callsign = sim['callsign_2'] if sim['callsign_1'] == flight_dict.get('callsign') else sim['callsign_1']
                level = sim['similarity_level']
                other_callsigns.append(f"{other_callsign}({level})")
            similarity_info = ', '.join(other_callsigns) if other_callsigns else "-"

            row_data = [
                row_num - 1,  # 번호
                flight_dict.get('callsign', ''),
                flight_dict.get('dept_airport_cd', ''),
                flight_dict.get('dest_airport_cd', ''),
                flight_dict.get('aircraft_type', ''),
                flight_dict.get('spd', ''),
                flight_dict.get('alt', ''),
                flight_dict.get('enr_minutes', 0),
                flight_dict.get('route', ''),
                flight_dict.get('eet', ''),
                flight_dict.get('eobt', ''),
                flight_dict.get('eobd', ''),
                sector_info,
                waypoint_info,
                similarity_info
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value if value else ""
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # 컬럼 너비 자동 조정
        column_widths = [8, 12, 12, 12, 12, 10, 10, 15, 25, 10, 12, 12, 30, 30, 20]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

        # 행 높이 설정
        ws.row_dimensions[1].height = 25
        for row in range(2, len(flights) + 2):
            ws.row_dimensions[row].height = 30

        # 파일을 메모리에 저장
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        # 파일 다운로드
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'flights_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )

    except Exception as e:
        logger.error(f"Excel 내보내기 오류: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'서버 오류: {str(e)}'
        }), 500


# ============================================================================
# 업로드 이력 엔드포인트
# ============================================================================

@app.route('/api/upload/history', methods=['GET'])
def get_upload_history():
    """
    업로드 이력 조회

    Query parameters:
        - limit: int (기본값: 20)

    Returns:
        JSON: {
            status: 'success',
            data: list of upload records
        }
    """
    try:
        limit = request.args.get('limit', 20, type=int)
        history = db_manager.get_upload_history(limit)

        return jsonify({
            'status': 'success',
            'data': [dict(record) for record in history]
        }), 200

    except Exception as e:
        logger.error(f"업로드 이력 조회 오류: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'서버 오류: {str(e)}'
        }), 500


# ============================================================================
# 에러 핸들러
# ============================================================================

@app.errorhandler(404)
def not_found(error):
    """404 에러 핸들러"""
    return jsonify({
        'status': 'error',
        'message': '요청한 리소스를 찾을 수 없습니다'
    }), 404


@app.errorhandler(500)
def internal_error(error):
    """500 에러 핸들러"""
    logger.error(f"서버 오류: {str(error)}")
    return jsonify({
        'status': 'error',
        'message': '서버 내부 오류가 발생했습니다'
    }), 500


# ============================================================================
# ⚡ CLI 빠른 시뮬레이션 엔드포인트
# ============================================================================

@app.route('/api/simulation/run-cli', methods=['POST'])
def run_simulation_cli():
    """
    CLI 방식의 빠른 시뮬레이션 (simulate_cli.py 기반)

    파일 업로드 + DB 저장 + 시뮬레이션까지 한 번에 처리
    """
    import subprocess
    import time
    import tempfile

    try:
        logger.info("CLI 시뮬레이션 시작...")
        logger.info(f"요청 메서드: {request.method}")
        logger.info(f"Content-Type: {request.content_type}")
        logger.info(f"Files: {request.files.keys()}")
        logger.info(f"Form: {request.form.keys()}")

        start_time = time.time()

        # 파일 확인
        if 'file' not in request.files:
            logger.error(f"파일 없음. 가능한 키: {list(request.files.keys())}")
            return jsonify({
                'status': 'error',
                'message': f'파일이 없습니다. 가능한 키: {list(request.files.keys())}'
            }), 400

        file = request.files['file']
        if file.filename == '':
            logger.error("파일명이 비어있음")
            return jsonify({
                'status': 'error',
                'message': '파일을 선택하세요'
            }), 400

        logger.info(f"파일명: {file.filename}")

        # 임시 파일로 저장
        with tempfile.NamedTemporaryFile(delete=False, suffix='.csv') as tmp:
            file.save(tmp.name)
            tmp_path = tmp.name

        try:
            # 데이터베이스 초기화 (필요한 경우)
            reset_db = request.form.get('reset_db', 'false').lower() == 'true'
            if reset_db:
                db_path = os.path.join(PROJECT_DIR, 'database', 'similarity_detector.db')
                if os.path.exists(db_path):
                    os.remove(db_path)
                    logger.info(f"데이터베이스 파일 삭제: {db_path}")

            # simulate_cli.py 실행
            cmd = [
                'python3',
                os.path.join(PROJECT_DIR, 'simulate_cli.py'),
                tmp_path
            ]

            logger.info(f"실행 명령: {' '.join(cmd)}")

            # 서브프로세스 실행 (타임아웃 없음 - 배경에서 실행)
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=3600  # 1시간 타임아웃
            )

            elapsed = time.time() - start_time

            if result.returncode != 0:
                logger.error(f"CLI 시뮬레이션 오류: {result.stderr}")
                return jsonify({
                    'status': 'error',
                    'message': f'시뮬레이션 실패: {result.stderr[:200]}'
                }), 500

            # DB에서 결과 조회
            stats = db_manager.get_statistics()

            logger.info(f"CLI 시뮬레이션 완료: {elapsed:.1f}초")

            return jsonify({
                'status': 'success',
                'message': f'{elapsed:.1f}초 완료',
                'data': {
                    'elapsed_time': elapsed,
                    'similarity_count': stats.get('total_similarities', 0),
                    'total_flights': stats.get('total_flights', 0),
                    'cross_sector_pairs': stats.get('cross_sector_overlap_pairs', 0),
                    'statistics': stats
                }
            }), 200

        finally:
            # 임시 파일 삭제
            if os.path.exists(tmp_path):
                os.remove(tmp_path)

    except subprocess.TimeoutExpired:
        logger.error("CLI 시뮬레이션 타임아웃")
        return jsonify({
            'status': 'error',
            'message': '시뮬레이션 타임아웃 (1시간 초과)'
        }), 504

    except Exception as e:
        logger.error(f"CLI 시뮬레이션 오류: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'서버 오류: {str(e)}'
        }), 500


@app.route('/api/simulation/reset-db', methods=['POST'])
def reset_db():
    """데이터베이스 초기화"""
    try:
        db_path = os.path.join(PROJECT_DIR, 'database', 'similarity_detector.db')

        if os.path.exists(db_path):
            os.remove(db_path)
            logger.info("데이터베이스 초기화 완료")

            # 새로운 DatabaseManager 인스턴스 생성
            global db_manager
            db_manager = DatabaseManager(db_path)

        return jsonify({
            'status': 'success',
            'message': '데이터베이스 초기화 완료'
        }), 200

    except Exception as e:
        logger.error(f"DB 초기화 오류: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'DB 초기화 실패: {str(e)}'
        }), 500


@app.route('/api/database/delete', methods=['POST'])
def delete_database():
    """
    데이터베이스 또는 특정 일자의 데이터 삭제

    Query parameters:
        - type: 'all' (전체 삭제) 또는 'date' (일자별 삭제)
        - date: type='date'일 때 삭제할 날짜 (YYYY-MM-DD 형식)
    """
    try:
        delete_type = request.args.get('type', 'all')
        date_str = request.args.get('date', None)

        logger.info(f"[삭제 요청] type={delete_type}, date={date_str}")
        print(f"[DEBUG] DELETE 요청: type={delete_type}, date={date_str}")

        if delete_type == 'all':
            # 전체 데이터베이스 삭제
            db_path = os.path.join(PROJECT_DIR, 'database', 'similarity_detector.db')

            if os.path.exists(db_path):
                os.remove(db_path)
                logger.info("데이터베이스 전체 삭제 완료")

                # 새로운 DatabaseManager 인스턴스 생성
                global db_manager
                db_manager = DatabaseManager(db_path)

            return jsonify({
                'status': 'success',
                'message': '전체 데이터베이스 삭제 완료'
            }), 200

        elif delete_type == 'date' and date_str:
            # 특정 일자의 데이터만 삭제
            try:
                # 날짜 형식 검증
                from datetime import datetime
                target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                date_str = str(target_date)
            except ValueError:
                return jsonify({
                    'status': 'error',
                    'message': '유효하지 않은 날짜 형식입니다 (YYYY-MM-DD)'
                }), 400

            # 해당 일자의 데이터 삭제
            # 1. similarities 테이블에서 해당 일자의 records 삭제
            # 2. flights 테이블에서 해당 일자의 records 삭제

            try:
                # 일자에 해당하는 flight id 조회
                flights_query = "SELECT id FROM flights WHERE eobd = ?"
                flights_result = db_manager.execute_query(flights_query, [date_str])
                flight_ids = [dict(row)['id'] for row in flights_result] if flights_result else []

                if flight_ids:
                    # flight id placeholder 생성
                    placeholders = ','.join('?' * len(flight_ids))

                    # sector_overlaps 삭제 (similarities를 통해)
                    delete_sector_overlaps = f"""
                        DELETE FROM sector_overlaps
                        WHERE similarity_id IN (
                            SELECT id FROM similarities
                            WHERE flight_id_1 IN ({placeholders}) OR flight_id_2 IN ({placeholders})
                        )
                    """
                    db_manager.execute_delete(delete_sector_overlaps, flight_ids + flight_ids)

                    # similarities 삭제
                    delete_similarities = f"""
                        DELETE FROM similarities
                        WHERE flight_id_1 IN ({placeholders}) OR flight_id_2 IN ({placeholders})
                    """
                    db_manager.execute_delete(delete_similarities, flight_ids + flight_ids)

                    # waypoint_times 삭제
                    delete_waypoints = f"""
                        DELETE FROM waypoint_times
                        WHERE flight_id IN ({placeholders})
                    """
                    db_manager.execute_delete(delete_waypoints, flight_ids)

                    # sector_times 삭제
                    delete_sectors = f"""
                        DELETE FROM sector_times
                        WHERE flight_id IN ({placeholders})
                    """
                    db_manager.execute_delete(delete_sectors, flight_ids)

                    # flights 삭제
                    delete_flights = f"""
                        DELETE FROM flights
                        WHERE id IN ({placeholders})
                    """
                    db_manager.execute_delete(delete_flights, flight_ids)

                logger.info(f"날짜 {date_str}의 데이터 삭제 완료 (항공편: {len(flight_ids)}개)")

                return jsonify({
                    'status': 'success',
                    'message': f'{date_str} 데이터 삭제 완료 ({len(flight_ids)}개 항공편)'
                }), 200

            except Exception as e:
                logger.error(f"일자별 데이터 삭제 오류: {str(e)}")
                return jsonify({
                    'status': 'error',
                    'message': f'데이터 삭제 실패: {str(e)}'
                }), 500
        else:
            return jsonify({
                'status': 'error',
                'message': '유효하지 않은 삭제 요청입니다'
            }), 400

    except Exception as e:
        logger.error(f"DB 삭제 오류: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'삭제 실패: {str(e)}'
        }), 500


# ============================================================================
# 항공기 기종 프로필 API (Aircraft Profiles)
# ============================================================================

@app.route('/api/aircraft-profiles', methods=['GET'])
def get_aircraft_profiles():
    """
    모든 항공기 기종 프로필 조회

    Returns:
        {
            'status': 'success',
            'data': [
                {
                    'icao_code': 'B77L',
                    'iata_code': '77L',
                    'manufacturer': 'Boeing',
                    'model': '777-200LRF',
                    'type_description': 'Long-range cargo freighter',
                    'default_speed_kmh': 905,
                    'default_climb_fpm': 2000,
                    'default_ceiling_fl': 430
                },
                ...
            ]
        }
    """
    try:
        profiles = db_manager.get_all_aircraft_profiles()
        return jsonify({
            'status': 'success',
            'data': profiles,
            'count': len(profiles)
        }), 200
    except Exception as e:
        logger.error(f"Failed to fetch aircraft profiles: {e}")
        return jsonify({
            'status': 'error',
            'message': f'항공기 기종 프로필 조회 실패: {str(e)}'
        }), 500


@app.route('/api/aircraft-profiles/<icao_code>', methods=['GET'])
def get_aircraft_profile(icao_code):
    """
    특정 항공기 기종 프로필 조회

    Args:
        icao_code: ICAO 항공기 코드 (e.g., B77L)

    Returns:
        {
            'status': 'success',
            'data': {...}
        }
    """
    try:
        profile = db_manager.get_aircraft_profile(icao_code.upper())
        if not profile:
            return jsonify({
                'status': 'error',
                'message': f'항공기 기종을 찾을 수 없습니다: {icao_code}'
            }), 404

        return jsonify({
            'status': 'success',
            'data': profile
        }), 200
    except Exception as e:
        logger.error(f"Failed to fetch aircraft profile {icao_code}: {e}")
        return jsonify({
            'status': 'error',
            'message': f'항공기 기종 프로필 조회 실패: {str(e)}'
        }), 500


@app.route('/api/aircraft-profiles', methods=['POST'])
def create_aircraft_profile():
    """
    새 항공기 기종 프로필 생성

    Request:
        {
            'icao_code': 'B77L',
            'iata_code': '77L',
            'manufacturer': 'Boeing',
            'model': '777-200LRF',
            'type_description': 'Long-range cargo freighter',
            'default_speed_kmh': 905,
            'default_speed_knots': 489,
            'default_climb_fpm': 2000,
            'default_ceiling_fl': 430,
            'notes': 'Optional notes'
        }

    Returns:
        {
            'status': 'success',
            'message': 'Aircraft profile created',
            'icao_code': 'B77L'
        }
    """
    try:
        data = request.get_json()
        if not data:
            return jsonify({
                'status': 'error',
                'message': '요청 본문이 비어있습니다'
            }), 400

        # 필수 필드 검증
        required_fields = ['icao_code', 'manufacturer', 'model', 'default_speed_kmh',
                          'default_speed_knots', 'default_climb_fpm', 'default_ceiling_fl']
        missing = [f for f in required_fields if f not in data]
        if missing:
            return jsonify({
                'status': 'error',
                'message': f'필수 필드가 누락되었습니다: {", ".join(missing)}'
            }), 400

        icao_code = data['icao_code'].upper()
        db_manager.insert_aircraft_profile(
            icao_code=icao_code,
            iata_code=data.get('iata_code', ''),
            manufacturer=data['manufacturer'],
            model=data['model'],
            type_description=data.get('type_description', ''),
            default_speed_kmh=data['default_speed_kmh'],
            default_speed_knots=data['default_speed_knots'],
            default_climb_fpm=data['default_climb_fpm'],
            default_ceiling_fl=data['default_ceiling_fl'],
            notes=data.get('notes')
        )

        return jsonify({
            'status': 'success',
            'message': '항공기 기종 프로필이 생성되었습니다',
            'icao_code': icao_code
        }), 201
    except Exception as e:
        logger.error(f"Failed to create aircraft profile: {e}")
        return jsonify({
            'status': 'error',
            'message': f'항공기 기종 프로필 생성 실패: {str(e)}'
        }), 500


@app.route('/api/aircraft-profiles/<icao_code>', methods=['PUT'])
def update_aircraft_profile(icao_code):
    """
    항공기 기종 프로필 업데이트

    Args:
        icao_code: ICAO 항공기 코드

    Request:
        {
            'default_speed_kmh': 910,
            'default_climb_fpm': 2100,
            ...
        }

    Returns:
        {
            'status': 'success',
            'message': 'Aircraft profile updated'
        }
    """
    try:
        data = request.get_json()
        if not data:
            return jsonify({
                'status': 'error',
                'message': '요청 본문이 비어있습니다'
            }), 400

        icao_code = icao_code.upper()

        # 기종 존재 확인
        existing = db_manager.get_aircraft_profile(icao_code)
        if not existing:
            return jsonify({
                'status': 'error',
                'message': f'항공기 기종을 찾을 수 없습니다: {icao_code}'
            }), 404

        # 업데이트 수행
        db_manager.update_aircraft_profile(icao_code, **data)

        return jsonify({
            'status': 'success',
            'message': '항공기 기종 프로필이 업데이트되었습니다'
        }), 200
    except ValueError as e:
        return jsonify({
            'status': 'error',
            'message': f'유효하지 않은 필드: {str(e)}'
        }), 400
    except Exception as e:
        logger.error(f"Failed to update aircraft profile {icao_code}: {e}")
        return jsonify({
            'status': 'error',
            'message': f'항공기 기종 프로필 업데이트 실패: {str(e)}'
        }), 500


@app.route('/api/aircraft-profiles/<icao_code>', methods=['DELETE'])
def delete_aircraft_profile(icao_code):
    """
    항공기 기종 프로필 삭제

    Args:
        icao_code: ICAO 항공기 코드

    Returns:
        {
            'status': 'success',
            'message': 'Aircraft profile deleted'
        }
    """
    try:
        icao_code = icao_code.upper()

        # 기종 존재 확인
        existing = db_manager.get_aircraft_profile(icao_code)
        if not existing:
            return jsonify({
                'status': 'error',
                'message': f'항공기 기종을 찾을 수 없습니다: {icao_code}'
            }), 404

        # 삭제 수행
        db_manager.delete_aircraft_profile(icao_code)

        return jsonify({
            'status': 'success',
            'message': '항공기 기종 프로필이 삭제되었습니다'
        }), 200
    except Exception as e:
        logger.error(f"Failed to delete aircraft profile {icao_code}: {e}")
        return jsonify({
            'status': 'error',
            'message': f'항공기 기종 프로필 삭제 실패: {str(e)}'
        }), 500


# ============================================================================
# 고도 상승 계산 비교 API (Climb Calculations)
# ============================================================================

@app.route('/api/flights/<int:flight_id>/climb-comparison', methods=['GET'])
def get_climb_comparison(flight_id):
    """
    항공편의 고도 상승 계산 방식 비교 결과

    Args:
        flight_id: 항공편 ID

    Returns:
        {
            'status': 'success',
            'flight_info': {
                'id': 1,
                'callsign': 'AAL123',
                'dept_airport': 'RKSI',
                'dest_airport': 'KSFO',
                'aircraft_type': 'B777',
                'calculated_speed_kmh': 905,
                'speed_source': 'aircraft_profile',
                'climb_rate_fpm': 2000,
                'cruise_flight_level': 430
            },
            'waypoints': [
                {
                    'sequence': 1,
                    'name': 'POINT1',
                    'simple_linear_time': '12:05:30',
                    'simple_linear_altitude_ft': 15000,
                    'eet_backtrack_time': '12:06:00',
                    'eet_backtrack_altitude_ft': 14800,
                    'time_difference_seconds': 30,
                    'altitude_difference_ft': 200
                },
                ...
            ],
            'statistics': {
                'total_waypoints': 15,
                'avg_time_diff_seconds': 45.5,
                'max_time_diff_seconds': 120,
                'avg_alt_diff_ft': 150.3,
                'max_alt_diff_ft': 500
            }
        }
    """
    try:
        # 항공편 정보 조회
        flight = db_manager.get_flight(flight_id)
        if not flight:
            return jsonify({
                'status': 'error',
                'message': f'항공편을 찾을 수 없습니다: {flight_id}'
            }), 404

        # 고도 상승 계산 결과 조회
        climb_data = db_manager.get_climb_calculations(flight_id)
        if not climb_data:
            return jsonify({
                'status': 'warning',
                'message': '해당 항공편의 고도 상승 계산 비교 데이터가 없습니다',
                'flight_info': dict(flight)
            }), 200

        # 통계 계산
        statistics = db_manager.get_climb_calculation_statistics(flight_id)

        return jsonify({
            'status': 'success',
            'flight_info': dict(flight),
            'waypoints': climb_data,
            'statistics': statistics if statistics else {}
        }), 200

    except Exception as e:
        logger.error(f"Failed to fetch climb comparison for flight {flight_id}: {e}")
        return jsonify({
            'status': 'error',
            'message': f'고도 상승 계산 비교 조회 실패: {str(e)}'
        }), 500


# ============================================================================
# 모델링 테스트 - 기종 고려 고도 상승 계산
# ============================================================================

@app.route('/api/test/calculate-flight', methods=['POST'])
def test_calculate_flight():
    """
    모델링 테스트 탭: 기종을 고려한 고도 상승 계산

    Request:
        {
            'callsign': 'KAL123',
            'aircraft_type': 'B77W',
            'speed': 'K0850',          -- CSV SPD (선택사항)
            'altitude': 'F370',        -- 순항 고도
            'dept': 'RKSI',            -- 출발지
            'dest': 'VHHH',            -- 도착지
            'eobd': '2025-12-26',      -- 날짜
            'eobt': '11:00',           -- 출발 시간
            'route': 'LAMEN A593 SADLI', -- 항로
            'eet': 'RKRR0015',         -- EET (선택사항)
            'info_cn': '...'           -- 추가 정보 (선택사항)
        }

    Returns:
        {
            'status': 'success',
            'flight_info': {
                'callsign': 'KAL123',
                'aircraft_type': 'B77W',
                'dept_airport': 'RKSI',
                'dest_airport': 'VHHH',
                'calculated_speed_kmh': 850,
                'speed_source': 'aircraft_profile',
                'climb_rate_fpm': 2000,
                'cruise_flight_level': 370,
                'departure_time': '11:00'
            },
            'waypoints': [
                {
                    'sequence': 1,
                    'name': 'LAMEN',
                    'estimated_time': '11:15',
                    'altitude_ft': 10000,
                    'is_climbing': true
                },
                ...
            ],
            'climb_analysis': {
                'method_a': {
                    'name': '단순 선형 상승',
                    'total_climb_minutes': 17.5,
                    'climb_distance_km': 87.5,
                    'cruise_distance_km': 412.5,
                    'total_time_minutes': 66.7
                },
                'method_b': {
                    'name': 'EET 역계산',
                    'total_climb_minutes': 16.8,
                    'climb_distance_km': 84.0
                }
            }
        }
    """
    try:
        data = request.get_json()
        if not data:
            return jsonify({
                'status': 'error',
                'message': '요청 데이터가 없습니다'
            }), 400

        # 입력 데이터 파싱
        callsign = data.get('callsign', 'TEST')
        aircraft_type = data.get('aircraft_type', '')
        speed_str = data.get('speed', '')
        altitude_str = data.get('altitude', 'F370')
        dept = data.get('dept', '')
        dest = data.get('dest', '')
        eobd = data.get('eobd', '')
        eobt = data.get('eobt', '00:00')
        route = data.get('route', '')

        # 고도 파싱
        from core.flight_processor import parse_altitude, get_aircraft_speed_and_climb, calculate_climb_time_simple, calculate_waypoints_with_eet
        from datetime import datetime, timedelta

        cruise_alt_ft = parse_altitude(altitude_str)
        if not cruise_alt_ft:
            cruise_alt_ft = 37000  # Default FL370

        # 기종별 속도 및 상승률 조회 (Days 1-5 구현 활용)
        aircraft_info = get_aircraft_speed_and_climb(db_manager, aircraft_type, speed_str)
        speed_kmh = aircraft_info['speed_kmh']
        climb_fpm = aircraft_info['climb_fpm']
        speed_source = aircraft_info['speed_source']

        # 항로 파싱 (공백/쉼표로 구분)
        waypoint_names = [w.strip() for w in route.replace(',', ' ').split() if w.strip()]
        if not waypoint_names:
            waypoint_names = ['DEP', 'DEST']

        # 항로 거리 계산 (세 가지 방식)
        total_distance_km = None
        expanded_waypoints = None

        # 1. 정확한 거리 계산 시도 - enroute.xlsx 좌표 사용 (CSV 업로드와 동일)
        try:
            from core import route_converter
            import pandas as pd
            import os

            # enroute.xlsx 로드 (프로젝트 루트/data/enroute/enroute.xlsx)
            # app.py: /Users/sein/Desktop/iccs/similarity_detector/portable_app/backend/app.py
            # 프로젝트 루트: /Users/sein/Desktop/iccs/similarity_detector
            base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            enroute_path = os.path.join(base_dir, 'data', 'enroute', 'enroute.xlsx')

            if os.path.exists(enroute_path):
                enroute_df, fix_col = route_converter.load_data(enroute_path)

                # 항로 자동 확장 (Y711 같은 유효하지 않은 waypoint → 실제 waypoint로 변환)
                # 먼저 사용자 입력을 모두 대문자로 변환
                route_waypoints = [w.upper() for w in waypoint_names]

                # 유효한 waypoint만 필터링
                valid_waypoints = [w for w in route_waypoints if w in enroute_df[fix_col].values]

                # 만약 유효한 waypoint가 2개 이상이면, 그 사이의 모든 waypoint를 자동으로 추가
                expanded_waypoints = []
                if len(valid_waypoints) >= 2:
                    # 첫 번째와 마지막 valid waypoint 사이의 모든 waypoint를 찾기
                    start_wp = valid_waypoints[0]
                    end_wp = valid_waypoints[-1]

                    # enroute.xlsx에서 두 waypoint 사이의 경로 찾기
                    all_waypoints_list = enroute_df[fix_col].unique().tolist()
                    try:
                        start_idx = all_waypoints_list.index(start_wp)
                        end_idx = all_waypoints_list.index(end_wp)

                        if start_idx <= end_idx:
                            expanded_waypoints = all_waypoints_list[start_idx:end_idx+1]
                        else:
                            expanded_waypoints = all_waypoints_list[end_idx:start_idx+1]

                        logger.info(f"[항로 확장] {start_wp}~{end_wp} 사이: {len(expanded_waypoints)}개 waypoint")
                    except (ValueError, IndexError) as e:
                        logger.debug(f"경로 확장 실패: {e}, airway 확장 시도")
                        route_str = ' '.join(waypoint_names)
                        expanded_waypoints = route_converter.expand_route(enroute_df, fix_col, route_str)
                else:
                    # 단일 또는 0개: airway 기반 확장 시도
                    route_str = ' '.join(waypoint_names)
                    expanded_waypoints = route_converter.expand_route(enroute_df, fix_col, route_str)

                logger.info(f"[항로 확장] 원본: {waypoint_names} ({len(waypoint_names)}개) → 확장: {len(expanded_waypoints)}개")

                # 확장된 waypoint로 필터링 (좌표가 있는 것만)
                korea_points = [p for p in expanded_waypoints if p in enroute_df[fix_col].values]
                logger.info(f"[필터링] 유효한 waypoint: {korea_points[:15] if len(korea_points) > 15 else korea_points}")

                # 좌표맵 생성
                coord_map = {}
                for _, row in enroute_df.iterrows():
                    fix = row[fix_col]
                    if fix not in coord_map:
                        coord_map[fix] = (row['LAT'], row['LON'])

                # 출발지 추가
                route_for_calc = []
                if dept:
                    route_for_calc.append(dept)
                route_for_calc.extend(korea_points if korea_points else waypoint_names)

                # Haversine으로 거리 계산
                from core.flight_processor import haversine
                total_distance_km = 0
                for i in range(len(route_for_calc) - 1):
                    current = route_for_calc[i]
                    next_wp = route_for_calc[i + 1]

                    if current in coord_map and next_wp in coord_map:
                        lat1, lon1 = coord_map[current]
                        lat2, lon2 = coord_map[next_wp]
                        leg_dist = haversine(lat1, lon1, lat2, lon2)
                        total_distance_km += leg_dist
                    else:
                        # 좌표 없으면 60km 추정
                        total_distance_km += 60

                logger.info(f"Calculated distance (enroute.xlsx): {total_distance_km:.1f}km for route {route_for_calc}")
        except Exception as e:
            logger.debug(f"enroute.xlsx distance calculation error: {e}")
            total_distance_km = None

        # 2. Fallback: waypoint_distance.py 사용
        if total_distance_km is None:
            try:
                from portable_app.core.waypoint_distance import calculate_route_distance
                route_for_calc = []
                if dept:
                    route_for_calc.append(dept)
                route_for_calc.extend(waypoint_names)

                total_distance_km = calculate_route_distance(route_for_calc)
                logger.info(f"Calculated distance (waypoint_distance.py): {total_distance_km:.1f}km for route {route_for_calc}")
            except Exception as e:
                logger.debug(f"waypoint_distance.py calculation error: {e}")
                total_distance_km = None

        # 3. Fallback: waypoint 개수 기반 추정
        if total_distance_km is None:
            logger.debug(f"Using estimation-based distance calculation")
            if len(waypoint_names) <= 1:
                total_distance_km = 300  # 최소 거리
            else:
                # 각 leg당 평균 60km (RKSI 근처 기준)
                total_distance_km = 60 * (len(waypoint_names) - 1)
                total_distance_km = max(100, min(total_distance_km, 1500))  # 범위 제한
        climb_result_a = calculate_climb_time_simple(
            total_distance_km, 0, cruise_alt_ft, climb_fpm, speed_kmh
        )

        # 출발 시간 파싱
        try:
            dept_time = datetime.strptime(f"{eobd} {eobt}", "%Y-%m-%d %H:%M")
        except:
            dept_time = datetime.now()

        # Method B: EET 역계산
        exit_time = dept_time + timedelta(
            minutes=climb_result_a['total_time_minutes']
        )

        # 확장된 waypoint로 데이터 생성 (실제 거리 기반)
        waypoints_to_use = korea_points if korea_points else waypoint_names
        logger.info(f"[Waypoint 계산] 사용할 waypoint: {len(waypoints_to_use)}개")

        points_data = []
        cumulative_dist = 0

        for i, name in enumerate(waypoints_to_use):
            if i == 0 and dept and dept in coord_map and name in coord_map:
                # 첫 번째: DEPT에서의 거리
                lat1, lon1 = coord_map[dept]
                lat2, lon2 = coord_map[name]
                leg_dist = haversine(lat1, lon1, lat2, lon2)
                cumulative_dist = leg_dist
            elif i > 0:
                # 이전 waypoint에서의 거리
                prev_name = waypoints_to_use[i-1]
                if prev_name in coord_map and name in coord_map:
                    lat1, lon1 = coord_map[prev_name]
                    lat2, lon2 = coord_map[name]
                    leg_dist = haversine(lat1, lon1, lat2, lon2)
                    cumulative_dist += leg_dist
                else:
                    # 좌표 없으면 균등 분배
                    cumulative_dist += (total_distance_km / max(len(waypoints_to_use) - 1, 1))
            else:
                cumulative_dist = 0

            points_data.append({
                'name': name,
                'dist': cumulative_dist
            })

        waypoints_b = calculate_waypoints_with_eet(
            exit_time, points_data, speed_kmh, 0, cruise_alt_ft, climb_fpm
        )

        # Method A로 waypoint 시간 계산
        waypoints_a = []
        current_time = dept_time
        for i, pt_data in enumerate(points_data):
            dist = pt_data['dist']
            if dist == 0:
                alt = 0
                is_clim = False
            else:
                climb_dist = climb_result_a['climb_distance_km']
                if dist <= climb_dist:
                    progress = dist / climb_dist if climb_dist > 0 else 0
                    alt = int(progress * cruise_alt_ft)
                    is_clim = True
                else:
                    alt = cruise_alt_ft
                    is_clim = False

            time_minutes = (dist / speed_kmh) * 60
            current_time = dept_time + timedelta(minutes=time_minutes)

            waypoints_a.append({
                'name': pt_data['name'],
                'time': current_time,
                'altitude_ft': alt,
                'is_climbing': is_clim
            })

        # 응답 데이터 구성
        flight_info = {
            'callsign': callsign,
            'aircraft_type': aircraft_type,
            'dept_airport': dept,
            'dest_airport': dest,
            'calculated_speed_kmh': speed_kmh,
            'speed_source': speed_source,
            'climb_rate_fpm': climb_fpm,
            'cruise_flight_level': int(cruise_alt_ft / 100) if cruise_alt_ft else 370,
            'departure_time': eobt
        }

        # Waypoint 응답 (Method A 기반)
        waypoints_response = []
        for i, wp in enumerate(waypoints_a):
            waypoints_response.append({
                'sequence': i + 1,
                'name': wp['name'],
                'estimated_time': wp['time'].strftime('%H:%M:%S') if isinstance(wp['time'], datetime) else wp['time'],
                'altitude_ft': wp['altitude_ft'],
                'is_climbing': wp['is_climbing']
            })

        # 고도 상승 분석 정보
        climb_analysis = {
            'method_a': {
                'name': '단순 선형 상승',
                'total_climb_minutes': round(climb_result_a['climb_time_minutes'], 1),
                'climb_distance_km': round(climb_result_a['climb_distance_km'], 1),
                'cruise_distance_km': round(climb_result_a['cruise_distance_km'], 1),
                'total_time_minutes': round(climb_result_a['total_time_minutes'], 1)
            },
            'method_b': {
                'name': 'EET 역계산',
                'total_climb_minutes': round(
                    (len(waypoints_b) - 1) if waypoints_b else 0, 1
                ),
                'climb_distance_km': round(total_distance_km * 0.2, 1)  # 근사값
            }
        }

        return jsonify({
            'status': 'success',
            'flight_info': flight_info,
            'waypoints': waypoints_response,
            'climb_analysis': climb_analysis
        }), 200

    except Exception as e:
        logger.error(f"Failed to calculate flight: {e}")
        return jsonify({
            'status': 'error',
            'message': f'계산 실패: {str(e)}'
        }), 500


# ============================================================================
# 메인
# ============================================================================

if __name__ == '__main__':
    logger.info(f"Flask 앱 시작... (포트: {API_PORT}, DEBUG: {DEBUG})")
    app.run(
        host='127.0.0.1',
        port=API_PORT,
        debug=DEBUG,
        use_reloader=True
    )
